"""
=============================================================================
Script 02 -- Nettoyage, enrichissement et detection IA
=============================================================================
Livrables :
  - data/openalex_clean.parquet
  - outputs/etape2_nettoyage.xlsx  (6 onglets)
=============================================================================
"""

import pandas as pd
import numpy as np
import json
import re
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

np.random.seed(42)

# =====================================================================
# MOTS-CLES IA
# =====================================================================

AI_KEYWORDS_DETAILED = {
    "Modeles et architectures": [
        (r"\bmachine learning\b", "machine learning"),
        (r"\bdeep learning\b", "deep learning"),
        (r"\bneural network\b", "neural network"),
        (r"\btransformer\b", "transformer"),
        (r"\bGPT\b", "GPT"),
        (r"\bBERT\b", "BERT"),
        (r"\bLLM\b", "LLM"),
        (r"\blarge language model\b", "large language model"),
        (r"\breinforcement learning\b", "reinforcement learning"),
        (r"\bconvolutional neural\b", "convolutional neural"),
        (r"\brecurrent neural\b", "recurrent neural"),
        (r"\bGAN\b", "GAN"),
        (r"\bgenerative adversarial\b", "generative adversarial"),
        (r"\bdiffusion model\b", "diffusion model"),
        (r"\bfoundation model\b", "foundation model"),
    ],
    "Techniques et methodes": [
        (r"\bpre-?trained model\b", "pre-trained model"),
        (r"\bfine-?tun\w+\b", "fine-tuning"),
        (r"\bprompt engineering\b", "prompt engineering"),
        (r"\bfew-?shot\b", "few-shot"),
        (r"\bzero-?shot\b", "zero-shot"),
        (r"\btransfer learning\b", "transfer learning"),
        (r"\bnatural language processing\b", "NLP"),
        (r"\bNLP\b", "NLP"),
        (r"\bcomputer vision\b", "computer vision"),
        (r"\btext mining\b", "text mining"),
        (r"\bsentiment analysis\b", "sentiment analysis"),
        (r"\bword embedding\b", "word embedding"),
        (r"\bknowledge graph\b", "knowledge graph"),
    ],
    "Outils et produits": [
        (r"\bChatGPT\b", "ChatGPT"),
        (r"\bGPT-4\b", "GPT-4"),
        (r"\bGPT-3\b", "GPT-3"),
        (r"\bClaude\b", "Claude"),
        (r"\bGemini\b", "Gemini"),
        (r"\bLLaMA\b", "LLaMA"),
        (r"\bCopilot\b", "Copilot"),
        (r"\bStable Diffusion\b", "Stable Diffusion"),
        (r"\bDALL-?E\b", "DALL-E"),
        (r"\bTensorFlow\b", "TensorFlow"),
        (r"\bPyTorch\b", "PyTorch"),
        (r"\bHugging\s*Face\b", "Hugging Face"),
        (r"\bOpenAI\b", "OpenAI"),
    ],
    "Concepts generaux": [
        (r"\bartificial intelligence\b", "artificial intelligence"),
        (r"\bgenerative AI\b", "generative AI"),
        (r"\bAI-assisted\b", "AI-assisted"),
        (r"\bAI-driven\b", "AI-driven"),
        (r"\bAI-powered\b", "AI-powered"),
        (r"\bimage recognition\b", "image recognition"),
    ],
}

# =====================================================================
# CLASSIFICATION GEOGRAPHIQUE
# =====================================================================

# Pays OCDE + hauts revenus hors-OCDE classiquement associes au Global North
# Source : membership OCDE 2024 + World Bank high-income classification
# Note : TR (Turquie) et MX (Mexique) conserves dans Global North (membres OCDE)
GLOBAL_NORTH_CODES = frozenset({
    "AU", "AT", "BE", "CA", "CL", "CO", "CZ", "DK", "EE", "FI",
    "FR", "DE", "GR", "HU", "IS", "IE", "IL", "IT", "JP", "KR",
    "LV", "LT", "LU", "MX", "NL", "NZ", "NO", "PL", "PT", "SK",
    "SI", "ES", "SE", "CH", "TR", "GB", "US",
    # Hauts revenus hors-OCDE (convention Global North en bibliometrie)
    "AE", "BH", "BN", "CY", "HK", "KW", "MT", "OM", "QA", "SA",
    "SG", "TW",
})


def parse_countries(countries_val):
    """Parse la colonne countries (liste JSON ou liste Python) en list de codes."""
    if countries_val is None:
        return []
    if isinstance(countries_val, list):
        return [str(c) for c in countries_val if c]
    if isinstance(countries_val, str):
        if not countries_val.strip():
            return []
        try:
            result = json.loads(countries_val)
            return [str(c) for c in result if c] if isinstance(result, list) else []
        except (json.JSONDecodeError, ValueError):
            return []
    return []


def classify_geo_zone(countries_val):
    """Classifie un article selon sa zone geographique dominante.

    Retourne :
      'Global North'  - tous les pays d'affiliation dans Global North
      'Global South'  - aucun pays dans Global North
      'International' - collaboration Nord-Sud
      'Unknown'       - pas de donnee pays
    """
    codes = set(parse_countries(countries_val))
    codes.discard("")
    if not codes:
        return "Unknown"
    has_north = bool(codes & GLOBAL_NORTH_CODES)
    has_south = bool(codes - GLOBAL_NORTH_CODES)
    if has_north and has_south:
        return "International"
    if has_north:
        return "Global North"
    return "Global South"


ALL_PATTERNS = []
for cat, kws in AI_KEYWORDS_DETAILED.items():
    for pattern, _ in kws:
        ALL_PATTERNS.append(pattern)
AI_PATTERN = re.compile("|".join(ALL_PATTERNS), re.IGNORECASE)


def detect_ai_mention(abstract):
    if not abstract or not isinstance(abstract, str):
        return False
    return bool(AI_PATTERN.search(abstract))


def classify_ai_intensity(abstract):
    if not abstract or not isinstance(abstract, str):
        return "none"
    matches = AI_PATTERN.findall(abstract)
    n = len(matches)
    if n == 0: return "none"
    elif n <= 2: return "peripheral"
    elif n <= 5: return "methodological"
    else: return "core"


# =====================================================================
# EXCEL STYLES (identiques au script 01)
# =====================================================================

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def write_header_row(ws, row, headers, color="2563EB"):
    fill = PatternFill("solid", fgColor=color)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_data_rows(ws, start_row, data_rows, ncols):
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ALT_FILL


def set_col_widths(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def write_title(ws, row, text, color="2563EB"):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=True, size=14, color=color)


# =====================================================================
# NETTOYAGE
# =====================================================================

def clean_data(df):
    cleaning_log = []

    cleaning_log.append(["0. Donnees brutes", "Chargement du fichier parquet", len(df), 0, len(df)])

    n_before = len(df)
    df = df.drop_duplicates(subset=["openalex_id"])
    cleaning_log.append(["1. Dedoublonnage", "Suppression doublons par openalex_id", n_before, n_before - len(df), len(df)])

    n_before = len(df)
    df = df[df["abstract"].str.len() > 50]
    cleaning_log.append(["2. Filtre abstract", "Abstract > 50 caracteres", n_before, n_before - len(df), len(df)])

    n_before = len(df)
    df = df[df["primary_discipline"] != "Unknown"]
    cleaning_log.append(["3. Filtre discipline", "Exclure discipline inconnue", n_before, n_before - len(df), len(df)])

    n_before = len(df)
    df["publication_date"] = pd.to_datetime(df["publication_date"], errors="coerce")
    df = df.dropna(subset=["publication_date"])
    cleaning_log.append(["4. Filtre date", "Exclure dates invalides", n_before, n_before - len(df), len(df)])

    # Enrichissement topics
    if "topics_json" in df.columns:
        df["topics_list"] = df["topics_json"].apply(
            lambda x: [t["display_name"] for t in json.loads(x)] if pd.notna(x) and x else []
        )

    df["quarter"] = df["publication_date"].dt.to_period("Q").astype(str)
    df["semester"] = df["year"].astype(str) + "-S" + ((df["publication_date"].dt.month - 1) // 6 + 1).astype(str)
    df["post_genai"] = (df["year"] >= 2023).astype(int)

    print("Classification geographique (geo_zone)...")
    df["geo_zone"] = df["countries"].apply(classify_geo_zone)
    cleaning_log.append(["6. Geo-zone", "Classification Global North/South/International", len(df), 0, len(df)])

    print("Detection des mentions IA dans les abstracts...")
    df["ai_mention"] = df["abstract"].apply(detect_ai_mention)
    df["ai_intensity"] = df["abstract"].apply(classify_ai_intensity)

    cleaning_log.append(["5. Enrichissement", "Ajout quarter, post_genai, ai_mention, ai_intensity", len(df), 0, len(df)])

    return df, cleaning_log


# =====================================================================
# EXCEL DE TRACABILITE
# =====================================================================

def create_cleaning_excel(df, cleaning_log, output_path):
    wb = Workbook()

    # --- Onglet 1 : Pipeline nettoyage ---
    ws1 = wb.active
    ws1.title = "1. Pipeline nettoyage"
    ws1.sheet_properties.tabColor = "2563EB"
    write_title(ws1, 1, "ETAPE 2 - NETTOYAGE ET PREPARATION DES DONNEES")
    ws1.cell(row=2, column=1, value="Date : " + datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws1["A2"].font = Font(name="Arial", size=10, italic=True, color="6B7280")

    headers = ["Etape", "Description", "Articles avant", "Articles supprimes", "Articles restants"]
    write_header_row(ws1, 4, headers)
    write_data_rows(ws1, 5, cleaning_log, 5)

    # Colorer en rouge les suppressions > 0
    for i, entry in enumerate(cleaning_log):
        cell = ws1.cell(row=5 + i, column=4)
        if cell.value and cell.value > 0:
            cell.font = Font(name="Arial", size=10, color="DC2626", bold=True)

    set_col_widths(ws1, {"A": 25, "B": 55, "C": 18, "D": 20, "E": 18})

    # --- Onglet 2 : Dictionnaire mots-cles IA ---
    ws2 = wb.create_sheet("2. Mots-cles IA")
    ws2.sheet_properties.tabColor = "DC2626"
    write_title(ws2, 1, "DICTIONNAIRE DE DETECTION IA")

    kw_headers = ["Categorie", "Mot-cle", "Regex"]
    write_header_row(ws2, 3, kw_headers, color="DC2626")

    kw_data = []
    for cat, keywords in AI_KEYWORDS_DETAILED.items():
        for pattern, label in keywords:
            kw_data.append([cat, label, pattern])
    write_data_rows(ws2, 4, kw_data, 3)

    # Note methodologique
    note_row = 4 + len(kw_data) + 1
    ws2.cell(row=note_row, column=1, value="Classification d'intensite :").font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=note_row+1, column=1, value="  none = 0 match | peripheral = 1-2 | methodological = 3-5 | core = 6+").font = Font(name="Arial", size=10)

    set_col_widths(ws2, {"A": 28, "B": 30, "C": 40})

    # --- Onglet 3 : IA par annee ---
    ws3 = wb.create_sheet("3. IA par annee")
    ws3.sheet_properties.tabColor = "059669"
    write_title(ws3, 1, "TAUX DE MENTION IA PAR ANNEE")

    yearly_ai = df.groupby("year").agg(
        n_total=("openalex_id", "count"),
        n_ai=("ai_mention", "sum"),
        n_peripheral=("ai_intensity", lambda x: (x == "peripheral").sum()),
        n_methodological=("ai_intensity", lambda x: (x == "methodological").sum()),
        n_core=("ai_intensity", lambda x: (x == "core").sum()),
    ).reset_index()
    yearly_ai["pct_ai"] = (yearly_ai["n_ai"] / yearly_ai["n_total"] * 100).round(1)

    h3 = ["Annee", "Total", "Mentions IA", "% IA", "Peripheral", "Methodological", "Core"]
    write_header_row(ws3, 3, h3, color="059669")

    d3 = []
    for _, r in yearly_ai.iterrows():
        d3.append([int(r["year"]), int(r["n_total"]), int(r["n_ai"]), r["pct_ai"],
                    int(r["n_peripheral"]), int(r["n_methodological"]), int(r["n_core"])])
    write_data_rows(ws3, 4, d3, 7)
    set_col_widths(ws3, {"A": 10, "B": 10, "C": 14, "D": 10, "E": 14, "F": 16, "G": 10})

    # --- Onglet 4 : IA par discipline ---
    ws4 = wb.create_sheet("4. IA par discipline")
    ws4.sheet_properties.tabColor = "7C3AED"
    write_title(ws4, 1, "TAUX DE MENTION IA PAR DISCIPLINE")

    disc_ai = df.groupby("primary_discipline").agg(
        n_total=("openalex_id", "count"),
        n_ai=("ai_mention", "sum"),
    ).reset_index()
    disc_ai["pct_ai"] = (disc_ai["n_ai"] / disc_ai["n_total"] * 100).round(1)
    disc_ai = disc_ai.sort_values("pct_ai", ascending=False)
    median_pct = disc_ai["pct_ai"].median()
    disc_ai["exposition"] = disc_ai["pct_ai"].apply(lambda x: "Forte" if x >= median_pct else "Faible")

    h4 = ["Discipline", "Total", "Mentions IA", "% IA", "Exposition"]
    write_header_row(ws4, 3, h4, color="7C3AED")

    d4 = []
    for _, r in disc_ai.iterrows():
        d4.append([r["primary_discipline"], int(r["n_total"]), int(r["n_ai"]), r["pct_ai"], r["exposition"]])
    write_data_rows(ws4, 4, d4, 5)

    note_row = 4 + len(d4) + 1
    ws4.cell(row=note_row, column=1, value="Seuil exposition : mediane = " + str(round(median_pct, 1)) + "%")
    ws4["A" + str(note_row)].font = Font(name="Arial", size=10, italic=True, color="6B7280")

    set_col_widths(ws4, {"A": 30, "B": 10, "C": 14, "D": 10, "E": 12})

    # --- Onglet 5 : Intensite IA ---
    ws5 = wb.create_sheet("5. Intensite IA")
    ws5.sheet_properties.tabColor = "F59E0B"
    write_title(ws5, 1, "DISTRIBUTION DE L'INTENSITE IA")

    intensity = df["ai_intensity"].value_counts()
    interp = {
        "none": "Aucune mention d'IA",
        "peripheral": "1-2 mentions, IA evoquee en passant",
        "methodological": "3-5 mentions, IA comme outil",
        "core": "6+ mentions, IA au coeur de l'article",
    }

    h5 = ["Niveau", "N articles", "% du total", "Interpretation"]
    write_header_row(ws5, 3, h5, color="F59E0B")

    d5 = []
    for level in ["none", "peripheral", "methodological", "core"]:
        count = int(intensity.get(level, 0))
        d5.append([level, count, round(count / len(df) * 100, 1), interp[level]])
    write_data_rows(ws5, 4, d5, 4)
    set_col_widths(ws5, {"A": 18, "B": 14, "C": 12, "D": 50})

    # --- Onglet 6 : Echantillon articles IA ---
    ws6 = wb.create_sheet("6. Echantillon articles IA")
    ws6.sheet_properties.tabColor = "EC4899"
    write_title(ws6, 1, "ECHANTILLON - 50 ARTICLES AVEC MENTION IA")

    ai_df = df[df["ai_mention"] == True]
    ai_sample = ai_df.sample(min(50, len(ai_df)), random_state=42) if len(ai_df) > 0 else ai_df.head(0)

    h6 = ["ID", "Titre", "Annee", "Discipline", "Intensite", "Citations", "Extrait abstract"]
    write_header_row(ws6, 3, h6, color="EC4899")

    d6 = []
    for _, r in ai_sample.iterrows():
        d6.append([
            str(r["openalex_id"])[-20:],
            str(r.get("title", ""))[:80],
            int(r["year"]),
            str(r["primary_discipline"]),
            str(r["ai_intensity"]),
            int(r["cited_by_count"]),
            str(r.get("abstract", ""))[:120],
        ])
    write_data_rows(ws6, 4, d6, 7)
    set_col_widths(ws6, {"A": 22, "B": 50, "C": 8, "D": 25, "E": 14, "F": 10, "G": 60})

    # --- Onglet 7 : Biais geographique ---
    ws7 = wb.create_sheet("7. Biais geographique")
    ws7.sheet_properties.tabColor = "059669"
    write_title(ws7, 1, "DISTRIBUTION PAR ZONE GEOGRAPHIQUE")
    ws7.cell(row=2, column=1, value="Global North = OCDE + hauts revenus | Global South = reste | International = co-publication Nord-Sud").font = Font(name="Arial", size=9, italic=True, color="6B7280")

    geo_yearly = df.groupby(["year", "geo_zone"]).size().unstack(fill_value=0).reset_index()
    for col in ["Global North", "Global South", "International", "Unknown"]:
        if col not in geo_yearly.columns:
            geo_yearly[col] = 0
    geo_yearly["total"] = geo_yearly[["Global North", "Global South", "International", "Unknown"]].sum(axis=1)
    for col in ["Global North", "Global South", "International", "Unknown"]:
        geo_yearly["pct_" + col.lower().replace(" ", "_")] = (geo_yearly[col] / geo_yearly["total"].clip(lower=1) * 100).round(1)

    h7 = ["Annee", "Global North", "Global South", "International", "Unknown",
          "% North", "% South", "% Intl", "% Unknown"]
    write_header_row(ws7, 4, h7, color="059669")
    d7 = []
    for _, r in geo_yearly.iterrows():
        d7.append([
            int(r["year"]),
            int(r["Global North"]), int(r["Global South"]),
            int(r["International"]), int(r["Unknown"]),
            r["pct_global_north"], r["pct_global_south"],
            r["pct_international"], r["pct_unknown"],
        ])
    write_data_rows(ws7, 5, d7, 9)
    set_col_widths(ws7, {"A": 10, "B": 14, "C": 14, "D": 14, "E": 12,
                          "F": 10, "G": 10, "H": 10, "I": 12})

    wb.save(output_path)
    print("Excel etape 2 : " + output_path)


# =====================================================================
# MAIN
# =====================================================================

EXTRACTION_DIR = os.path.join("outputs", "etape1_extraction")


def rebuild_raw_parquet(raw_path):
    """Consolide tous les CSV annuels en un seul fichier parquet brut."""
    csv_files = sorted([
        os.path.join(EXTRACTION_DIR, f)
        for f in os.listdir(EXTRACTION_DIR)
        if f.startswith("openalex_") and f.endswith(".csv")
    ])
    if not csv_files:
        return False
    print("Consolidation de " + str(len(csv_files)) + " fichiers CSV annuels...")
    chunks = []
    for path in csv_files:
        df_year = pd.read_csv(path, low_memory=False)
        print("  " + os.path.basename(path) + " : " + str(len(df_year)) + " lignes")
        chunks.append(df_year)
    df_all = pd.concat(chunks, ignore_index=True)
    print("Total consolide : " + str(len(df_all)) + " articles")
    os.makedirs(os.path.dirname(raw_path), exist_ok=True)
    df_all.to_parquet(raw_path, index=False)
    print("Parquet brut sauvegarde : " + raw_path)
    return True


def main():
    raw_path = "data/openalex_raw.parquet"
    clean_path = "data/openalex_clean.parquet"
    excel_path = "outputs/etape2_nettoyage.xlsx"

    print("=" * 70)
    print("  ETAPE 2 -- NETTOYAGE ET DETECTION IA")
    print("=" * 70)

    # Reconstruction du parquet brut depuis les CSV annuels si necessaire
    csv_available = os.path.isdir(EXTRACTION_DIR) and any(
        f.startswith("openalex_") and f.endswith(".csv")
        for f in os.listdir(EXTRACTION_DIR)
    )
    if csv_available:
        rebuild_raw_parquet(raw_path)
    elif not os.path.exists(raw_path):
        print("ERREUR : " + raw_path + " introuvable.")
        print("  Lance d'abord : python 01_extract_openalex.py --api-key <CLE>")
        return

    print("Chargement des donnees brutes...")
    df = pd.read_parquet(raw_path)
    print("  " + str(len(df)) + " articles charges")

    df, cleaning_log = clean_data(df)

    os.makedirs(os.path.dirname(clean_path), exist_ok=True)
    df.to_parquet(clean_path, index=False)
    print("data : " + clean_path + " (" + str(len(df)) + " articles)")

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    create_cleaning_excel(df, cleaning_log, excel_path)

    print()
    print("Resume :")
    print("  Disciplines : " + str(df["primary_discipline"].nunique()))
    print("  Mentions IA : " + str(int(df["ai_mention"].sum())) + " (" + str(round(df["ai_mention"].mean() * 100, 1)) + "%)")
    print()
    print("Etape 2 terminee. Prochaine commande :")
    print("  python 03_descriptive_analysis.py")


if __name__ == "__main__":
    main()