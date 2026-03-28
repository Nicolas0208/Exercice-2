"""
=============================================================================
Script 01 -- Extraction des donnees OpenAlex
=============================================================================
Livrables produits :
  - outputs/etape1_extraction/openalex_<annee>.xlsx  (un fichier par annee)
  - outputs/etape1_extraction/etape1_log.xlsx         (trace globale)

Dependances :
  pip install requests pandas openpyxl

Volume cible : ~150 000 articles/an => ~1 650 000 articles au total
Strategie    : 15 seeds x sample=10000 x 50 pages de 200 = 150 000 par an
=============================================================================
"""

import re
import requests
import pandas as pd
import time
import json
import argparse
import os
import logging
from datetime import datetime

# Caractères interdits dans les cellules Excel (openpyxl)
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def clean_str(value):
    """Supprime les caractères de contrôle illégaux pour Excel."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub("", value)
    return value

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_URL       = "https://api.openalex.org/works"
YEARS          = range(2015, 2026)
PER_PAGE       = 200
PAGES_PER_SEED = 50          # 50 pages x 200 = 10 000 articles par seed
SEEDS_PER_YEAR = 15          # 15 seeds => ~150 000 articles par an
SAMPLE_SIZE    = 10_000      # sample= par requete seed
RATE_LIMIT_DELAY = 0.1       # secondes entre requetes
OUTPUT_DIR     = os.path.join("outputs", "etape1_extraction")

COST_PER_REQUEST = 0.0001    # $ par requete (estimation)

SELECT_FIELDS = ",".join([
    "id", "title", "publication_date", "type", "cited_by_count",
    "concepts", "topics", "abstract_inverted_index",
    "authorships", "primary_location", "ids",
])

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S,%f"[:-3],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers de parsing
# ---------------------------------------------------------------------------

def reconstruct_abstract(inverted_index):
    if not inverted_index:
        return ""
    positions = []
    for word, pos_list in inverted_index.items():
        for pos in pos_list:
            positions.append((pos, word))
    positions.sort()
    return " ".join(w for _, w in positions)


def extract_concepts(concepts, max_level=2):
    if not concepts:
        return []
    return [
        {
            "id": c.get("id", ""),
            "display_name": c.get("display_name", ""),
            "level": c.get("level", -1),
            "score": c.get("score", 0.0),
        }
        for c in concepts if c.get("level", 99) <= max_level
    ]


def extract_topics(topics):
    if not topics:
        return []
    return [
        {
            "id": t.get("id", ""),
            "display_name": t.get("display_name", ""),
            "subfield": t.get("subfield", {}).get("display_name", ""),
            "field": t.get("field", {}).get("display_name", ""),
            "domain": t.get("domain", {}).get("display_name", ""),
            "score": t.get("score", 0.0),
        }
        for t in topics
    ]


def parse_work(work, year):
    abstract  = reconstruct_abstract(work.get("abstract_inverted_index", {}))
    concepts  = extract_concepts(work.get("concepts", []))
    topics    = extract_topics(work.get("topics", []))

    l0 = [c for c in concepts if c["level"] == 0]
    primary_discipline = l0[0]["display_name"] if l0 else "Unknown"

    source = ""
    loc = work.get("primary_location") or {}
    src = loc.get("source") or {}
    source = src.get("display_name", "")

    countries = []
    for auth in work.get("authorships", []):
        for inst in auth.get("institutions", []):
            cc = inst.get("country_code", "")
            if cc:
                countries.append(cc)

    return {
        "openalex_id":        work.get("id", ""),
        "title":              clean_str(work.get("title", "")),
        "abstract":           clean_str(abstract),
        "publication_date":   work.get("publication_date", ""),
        "year":               year,
        "cited_by_count":     work.get("cited_by_count", 0),
        "n_authors":          len(work.get("authorships", [])),
        "primary_discipline": primary_discipline,
        "source_journal":     source,
        "countries":          json.dumps(countries),
        "concepts_json":      json.dumps(concepts),
        "topics_json":        json.dumps(topics),
        "n_concepts":         len(concepts),
        "n_topics":           len(topics),
    }

# ---------------------------------------------------------------------------
# Extraction d'un seed (jusqu'a PAGES_PER_SEED pages)
# ---------------------------------------------------------------------------

def fetch_seed(year, seed, api_key, seen_ids, request_counter):
    """Recupere jusqu'a PAGES_PER_SEED*PER_PAGE articles pour un seed donne.
    Utilise la pagination par page= (et non cursor=) car OpenAlex ne renvoie
    pas de next_cursor quand sample= est actif.
    Retourne (records_nouveaux, n_doublons, n_requetes_effectuees).
    """
    headers = {"Authorization": f"Bearer {api_key}"}

    new_records = []
    n_doublons  = 0
    n_req       = 0

    for page_num in range(1, PAGES_PER_SEED + 1):
        params = {
            "filter":   f"publication_year:{year},type:article,has_abstract:true",
            "select":   SELECT_FIELDS,
            "per_page": PER_PAGE,
            "page":     page_num,
            "sample":   SAMPLE_SIZE,
            "seed":     seed,
        }

        # Tentatives avec retry
        for attempt in range(3):
            try:
                resp = requests.get(BASE_URL, params=params, headers=headers, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception as e:
                wait = 5 * (attempt + 1)
                log.warning(f"    Erreur page {page_num} seed {seed} : {e}. Retry dans {wait}s...")
                time.sleep(wait)
        else:
            log.error(f"    Abandon seed {seed} page {page_num} apres 3 tentatives.")
            break

        n_req += 1
        request_counter[0] += 1

        results = data.get("results", [])
        if not results:
            break

        for work in results:
            wid = work.get("id", "")
            if wid in seen_ids:
                n_doublons += 1
                continue
            seen_ids.add(wid)
            new_records.append(parse_work(work, year))

        time.sleep(RATE_LIMIT_DELAY)

    return new_records, n_doublons, n_req

# ---------------------------------------------------------------------------
# Extraction d'une annee complete
# ---------------------------------------------------------------------------

def fetch_year(year, api_key, total_request_counter, global_total, global_doublons):
    """Extrait tous les seeds pour une annee. Retourne la liste de records."""
    log.info(f"--- Annee {year} ---")

    seen_ids    = set()   # reinitialise par annee pour economiser la RAM
    year_records = []
    year_doublons = 0

    seeds_done = 0
    for seed in range(1, SEEDS_PER_YEAR + 1):
        records, n_doublons, n_req = fetch_seed(
            year, seed, api_key, seen_ids, total_request_counter
        )

        year_records.extend(records)
        year_doublons += n_doublons
        global_doublons[0] += n_doublons
        global_total[0] += len(records)
        seeds_done += 1

        cost = total_request_counter[0] * COST_PER_REQUEST
        req_num = total_request_counter[0]

        log.info(
            f"  [{req_num}] Annee {year} | seed {seed} | "
            f"batch={len(records) + n_doublons} | "
            f"nouveaux={len(records)} | "
            f"total articles={global_total[0]} | "
            f"doublons={global_doublons[0]} | "
            f"requetes={req_num} | "
            f"cout estime={cost:.4f} $"
        )

    log.info(
        f"  Bilan annee {year} : {seeds_done} seeds | "
        f"{len(year_records)} articles uniques | "
        f"{year_doublons} doublons"
    )
    return year_records

# ---------------------------------------------------------------------------
# Export Excel par annee (openpyxl)
# ---------------------------------------------------------------------------

COLS = [
    "openalex_id", "title", "abstract", "publication_date", "year",
    "cited_by_count", "n_authors", "primary_discipline", "source_journal",
    "countries", "concepts_json", "topics_json", "n_concepts", "n_topics",
]


_ILLEGAL_XML_CHARS = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ufffe\uffff]"
)

def _clean_str(val):
    if isinstance(val, str):
        # Step 1: drop lone surrogates via encode/decode (re can't match them reliably)
        val = val.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
        # Step 2: remove remaining illegal XML 1.0 characters
        val = _ILLEGAL_XML_CHARS.sub("", val)
    return val

def save_year_excel(records, year, output_dir):
    """Convertit la liste en DataFrame, sauvegarde en XLSX, libere la memoire."""
    df = pd.DataFrame(records, columns=COLS)
    # df.map is element-wise only in pandas >= 2.1; use per-column apply for compatibility
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].apply(lambda x: _clean_str(x) if isinstance(x, str) else x)
    n = len(df)

    # CSV backup — always saved first so data is never lost on Excel failure
    csv_path = os.path.join(output_dir, f"openalex_{year}.csv")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info(f"  => CSV backup : {csv_path} ({n} articles)")

    path = os.path.join(output_dir, f"openalex_{year}.xlsx")
    try:
        df.to_excel(path, index=False, engine="openpyxl")
        log.info(f"  => Sauvegarde : {path} ({n} articles)")
    except Exception as exc:
        log.warning(f"  Echec sauvegarde XLSX ({exc}). Donnees disponibles dans {csv_path}")

    del df
    return n

# ---------------------------------------------------------------------------
# Log global Excel (openpyxl)
# ---------------------------------------------------------------------------

def save_log_excel(yearly_stats, output_dir):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="2563EB")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="center", vertical="center")
    alt_fill   = PatternFill("solid", fgColor="F3F4F6")

    wb = Workbook()
    ws = wb.active
    ws.title = "Log extraction"
    ws.sheet_properties.tabColor = "059669"

    ws.cell(row=1, column=1, value="ETAPE 1 - LOG D'EXTRACTION OPENALEX").font = Font(
        name="Arial", bold=True, size=14, color="2563EB"
    )
    ws.cell(row=2, column=1, value=f"Date : {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Arial", size=10, italic=True, color="6B7280"
    )

    headers = ["Annee", "Seeds", "Articles uniques", "Doublons", "Requetes", "Duree (sec)", "Fichier"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin

    total_articles = 0
    for i, stat in enumerate(yearly_stats):
        row = 5 + i
        values = [
            stat["year"], stat["seeds"], stat["articles"],
            stat["doublons"], stat["requetes"],
            round(stat["duration_sec"], 1),
            f"openalex_{stat['year']}.xlsx",
        ]
        total_articles += stat["articles"]
        for j, val in enumerate(values, 1):
            c = ws.cell(row=row, column=j, value=val)
            c.font = data_font
            c.alignment = data_align
            c.border = thin
            if i % 2 == 1:
                c.fill = alt_fill

    # Ligne total
    tr = 5 + len(yearly_stats)
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=tr, column=3, value=total_articles).font = Font(name="Arial", bold=True, size=11)

    for col, w in zip("ABCDEFG", [10, 10, 18, 12, 12, 14, 30]):
        ws.column_dimensions[col].width = w

    path = os.path.join(output_dir, "etape1_log.xlsx")
    wb.save(path)
    log.info(f"Log global : {path}")

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Extraction OpenAlex - volume x150")
    parser.add_argument("--api-key", required=True, help="Clé API OpenAlex")
    parser.add_argument("--output-dir", default=OUTPUT_DIR)
    parser.add_argument("--seeds-per-year", type=int, default=SEEDS_PER_YEAR)
    parser.add_argument("--pages-per-seed", type=int, default=PAGES_PER_SEED)
    parser.add_argument("--year", type=int, default=None,
                        help="Traiter une seule année (ex: 2015). Si absent, toutes les années.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    years = [args.year] if args.year else list(YEARS)
    n_years       = len(years)
    total_seeds   = args.seeds_per_year * n_years
    total_req_est = total_seeds * args.pages_per_seed
    articles_max  = total_req_est * PER_PAGE
    cost_est      = total_req_est * COST_PER_REQUEST

    log.info("=" * 60)
    log.info(f"OpenAlex Extractor — démarrage {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"Années         : {min(years)} → {max(years)} ({n_years} ans)")
    log.info(f"Seeds / année  : {args.seeds_per_year}")
    log.info(f"Pages / seed   : {args.pages_per_seed}  (x{PER_PAGE} = {args.pages_per_seed*PER_PAGE} articles max/seed)")
    log.info(f"Requêtes prévues : {total_req_est}  (~{cost_est:.2f} $)")
    log.info(f"Articles max   : {articles_max:,}  (si 0 doublons)")
    log.info(f"Sortie         : {args.output_dir}/openalex_<annee>.xlsx")
    log.info(f"Dépendances    : pip install requests pandas openpyxl")
    log.info("=" * 60)

    total_request_counter = [0]   # liste pour passage par reference
    global_total          = [0]
    global_doublons       = [0]
    yearly_stats          = []

    for year in years:
        t0 = time.time()

        records = fetch_year(
            year, args.api_key,
            total_request_counter, global_total, global_doublons
        )

        n_saved = save_year_excel(records, year, args.output_dir)

        elapsed = time.time() - t0
        yearly_stats.append({
            "year":        year,
            "seeds":       args.seeds_per_year,
            "articles":    n_saved,
            "doublons":    global_doublons[0],
            "requetes":    total_request_counter[0],
            "duration_sec": elapsed,
        })

        # Liberation memoire avant l'annee suivante
        del records

    save_log_excel(yearly_stats, args.output_dir)

    total_articles = sum(s["articles"] for s in yearly_stats)
    cost_final     = total_request_counter[0] * COST_PER_REQUEST

    log.info("=" * 60)
    log.info("EXTRACTION TERMINÉE")
    log.info(f"Requêtes effectuées : {total_request_counter[0]}")
    log.info(f"Articles uniques    : {total_articles:,}")
    log.info(f"Doublons ignorés    : {global_doublons[0]}")
    log.info(f"Coût estimé         : {cost_final:.4f} $")
    log.info(f"Fichiers de sortie  : {args.output_dir}/openalex_<annee>.xlsx")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
