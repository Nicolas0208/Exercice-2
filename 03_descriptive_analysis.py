"""
=============================================================================
Script 03 -- Analyse descriptive complete
=============================================================================
Livrables :
  - outputs/etape3_analyse_descriptive.xlsx  (10 onglets)
  - outputs/figures/fig01 a fig07            (7 PNG)
=============================================================================
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
from itertools import combinations
from scipy import stats
import os
import json
import warnings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

plt.rcParams.update({
    "figure.dpi": 150, "figure.figsize": (12, 6), "font.size": 11,
    "font.family": "sans-serif", "axes.spines.top": False,
    "axes.spines.right": False, "axes.grid": True, "grid.alpha": 0.3,
})

COLORS = {"primary": "#2563EB", "secondary": "#DC2626", "accent": "#059669",
          "gray": "#6B7280", "light_blue": "#DBEAFE", "light_red": "#FEE2E2"}

FIG_DIR = "outputs/figures"
os.makedirs(FIG_DIR, exist_ok=True)


def normalize_topics(topics):
    if topics is None:
        return []
    if hasattr(topics, "tolist"):
        return [str(t) for t in topics.tolist()]
    if isinstance(topics, list):
        return [str(t) for t in topics]
    if isinstance(topics, str):
        try:
            return [str(t) for t in json.loads(topics)]
        except:
            return []
    return []


# =====================================================================
# PARSERS COLONNES STRUCTUREES
# =====================================================================

def parse_countries_for_cooc(val):
    """Parse la colonne countries (JSON ou liste) en liste de codes ISO."""
    if val is None:
        return []
    if isinstance(val, list):
        return [str(c) for c in val if c]
    if isinstance(val, str):
        if not val.strip():
            return []
        try:
            result = json.loads(val)
            return [str(c) for c in result if c] if isinstance(result, list) else []
        except (json.JSONDecodeError, ValueError):
            return []
    return []


def parse_concepts(val):
    """Parse la colonne concepts_json en liste de noms de concepts."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return []
    if isinstance(val, list):
        names = []
        for item in val:
            if isinstance(item, dict):
                n = item.get("display_name") or item.get("name") or ""
                if n:
                    names.append(str(n))
            elif item:
                names.append(str(item))
        return names
    if isinstance(val, str):
        if not val.strip():
            return []
        try:
            result = json.loads(val)
            if not isinstance(result, list):
                return []
            names = []
            for item in result:
                if isinstance(item, dict):
                    n = item.get("display_name") or item.get("name") or ""
                    if n:
                        names.append(str(n))
                elif item:
                    names.append(str(item))
            return names
        except (json.JSONDecodeError, ValueError):
            return []
    return []


# =====================================================================
# METRIQUES DE DIVERSITE
# =====================================================================

def shannon_entropy(counts):
    c = np.array(counts, dtype=float)
    c = c[c > 0]
    t = c.sum()
    if t == 0: return 0
    p = c / t
    return -np.sum(p * np.log2(p))

def simpson_diversity(counts):
    c = np.array(counts, dtype=float)
    t = c.sum()
    if t <= 1: return 0
    return 1 - np.sum(c * (c - 1)) / (t * (t - 1))

def effective_n(counts):
    return 2 ** shannon_entropy(counts)

def gini_coefficient(counts):
    c = np.sort(np.array(counts, dtype=float))
    n = len(c)
    if n == 0 or c.sum() == 0: return 0
    idx = np.arange(1, n + 1)
    return (2 * np.sum(idx * c) - (n + 1) * c.sum()) / (n * c.sum())

def hhi_index(counts):
    c = np.array(counts, dtype=float)
    t = c.sum()
    if t == 0: return 0
    return np.sum((c / t) ** 2)

def top_n_share(counts, n=5):
    c = np.sort(np.array(counts, dtype=float))[::-1]
    t = c.sum()
    if t == 0: return 0
    return c[:n].sum() / t


# =====================================================================
# EXCEL STYLES
# =====================================================================

THIN_BORDER = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")

def write_header_row(ws, row, headers, color="2563EB"):
    fill = PatternFill("solid", fgColor=color)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def write_data_rows(ws, start_row, data_rows, ncols):
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ALT_FILL

def set_col_widths(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

def write_title(ws, row, text, color="2563EB"):
    ws.cell(row=row, column=1, value=text).font = Font(name="Arial", bold=True, size=14, color=color)


# =====================================================================
# CALCUL METRIQUES PAR PERIODE
# =====================================================================

def compute_period_metrics(df, period_col):
    results = []
    for period, group in df.groupby(period_col):
        all_topics = []
        for topics in group["topics_list"]:
            all_topics.extend(normalize_topics(topics))
        counter = Counter(all_topics)
        c = list(counter.values())
        results.append({
            "period": str(period),
            "n_articles": len(group),
            "n_ai_articles": int(group["ai_mention"].sum()),
            "pct_ai": round(group["ai_mention"].mean() * 100, 1),
            "n_unique_topics": len(counter),
            "shannon": round(shannon_entropy(c), 4),
            "simpson": round(simpson_diversity(c), 4),
            "effective_n": round(effective_n(c), 1),
            "gini": round(gini_coefficient(np.array(c)), 4),
            "hhi": round(hhi_index(c), 6),
            "top5_share": round(top_n_share(c, 5) * 100, 2),
            "top10_share": round(top_n_share(c, 10) * 100, 2),
        })
    return pd.DataFrame(results).sort_values("period").reset_index(drop=True)


# =====================================================================
# FIGURES
# =====================================================================

def fig01_volume(df):
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    yearly = df.groupby("year").agg(n=("openalex_id", "count"), n_ai=("ai_mention", "sum")).reset_index()
    yearly["pct_ai"] = yearly["n_ai"] / yearly["n"] * 100

    ax = axes[0]
    ax.bar(yearly["year"], yearly["n"], color=COLORS["light_blue"], edgecolor=COLORS["primary"], label="Total")
    ax.bar(yearly["year"], yearly["n_ai"], color=COLORS["primary"], alpha=0.7, label="Mention IA")
    ax2 = ax.twinx()
    ax2.plot(yearly["year"], yearly["pct_ai"], "o-", color=COLORS["secondary"], lw=2, ms=6, label="% IA")
    ax2.set_ylabel("% mention IA", color=COLORS["secondary"])
    ax2.spines["right"].set_visible(True)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee"); ax.set_ylabel("N articles")
    ax.set_title("(a) Volume et adoption IA")
    h1, l1 = ax.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax.legend(h1 + h2, l1 + l2, loc="upper left", fontsize=9)

    top = df["primary_discipline"].value_counts().head(10)
    ax = axes[1]
    ax.barh(range(len(top)), top.values, color=plt.cm.Set3(np.linspace(0, 1, len(top))))
    ax.set_yticks(range(len(top))); ax.set_yticklabels(top.index)
    ax.set_xlabel("N articles"); ax.set_title("(b) Top 10 disciplines"); ax.invert_yaxis()

    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig01_volume_composition.png", bbox_inches="tight")
    plt.close()
    print("  Figure 1 : Volume et composition")


def fig02_diversity(metrics_q):
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    x = list(range(len(metrics_q)))
    quarters = metrics_q["period"].tolist()
    ts = max(1, len(quarters) // 12)

    ci = None
    for i, q in enumerate(quarters):
        if q >= "2023Q1":
            ci = i
            break

    def vline(ax):
        if ci is not None:
            ax.axvline(ci, color=COLORS["gray"], ls="--", alpha=0.5)

    # (a) Shannon
    ax = axes[0, 0]
    ax.plot(x, metrics_q["shannon"], "-o", color=COLORS["primary"], ms=3, lw=1.5)
    if len(x) > 2:
        z = np.polyfit(x, metrics_q["shannon"].values, 2)
        ax.plot(x, np.poly1d(z)(x), "--", color=COLORS["secondary"], alpha=0.7, label="Tendance")
        ax.legend(fontsize=9)
    ax.set_title("(a) Entropie de Shannon"); ax.set_ylabel("Bits")
    ax.set_xticks(x[::ts]); ax.set_xticklabels([quarters[i] for i in x[::ts]], rotation=45, fontsize=8)
    vline(ax)

    # (b) Effective N
    ax = axes[0, 1]
    ax.plot(x, metrics_q["effective_n"], "-o", color=COLORS["accent"], ms=3, lw=1.5)
    ax.set_title("(b) Nombre effectif de topics"); ax.set_ylabel("2^H")
    ax.set_xticks(x[::ts]); ax.set_xticklabels([quarters[i] for i in x[::ts]], rotation=45, fontsize=8)
    vline(ax)

    # (c) Gini
    ax = axes[1, 0]
    ax.plot(x, metrics_q["gini"], "-o", color=COLORS["secondary"], ms=3, lw=1.5)
    ax.set_title("(c) Coefficient de Gini"); ax.set_ylabel("Gini")
    ax.set_xticks(x[::ts]); ax.set_xticklabels([quarters[i] for i in x[::ts]], rotation=45, fontsize=8)
    vline(ax)

    # (d) Top-5 share
    ax = axes[1, 1]
    ax.plot(x, metrics_q["top5_share"], "-o", color="#7C3AED", ms=3, lw=1.5)
    ax.set_title("(d) Part des 5 topics dominants"); ax.set_ylabel("% du total")
    ax.set_xticks(x[::ts]); ax.set_xticklabels([quarters[i] for i in x[::ts]], rotation=45, fontsize=8)
    vline(ax)

    plt.suptitle("Evolution de la diversite thematique (2015-2026)", fontsize=14, fontweight="bold", y=1.01)
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig02_diversite_temporelle.png", bbox_inches="tight")
    plt.close()
    print("  Figure 2 : Series temporelles de diversite")


def fig03_heatmap(df):
    topic_year = {}
    for _, row in df.iterrows():
        year = row["year"]
        for t in normalize_topics(row["topics_list"]):
            key = (t, year)
            topic_year[key] = topic_year.get(key, 0) + 1
    records = [{"topic": k[0], "year": k[1], "count": v} for k, v in topic_year.items()]
    pdf = pd.DataFrame(records)
    if pdf.empty:
        print("  Figure 3 : Pas assez de donnees")
        return

    top = pdf.groupby("topic")["count"].sum().nlargest(min(25, len(pdf["topic"].unique()))).index.tolist()
    matrix = pdf[pdf["topic"].isin(top)].pivot_table(index="topic", columns="year", values="count", fill_value=0)
    matrix_pct = matrix.div(matrix.sum(axis=0), axis=1) * 100

    fig, ax = plt.subplots(figsize=(16, max(6, len(top) * 0.4)))
    sns.heatmap(matrix_pct, cmap="YlOrRd", linewidths=0.5, ax=ax, cbar_kws={"label": "% du total"})
    ax.set_title("Part relative des topics principaux par annee (%)", fontsize=13)
    if 2023 in matrix_pct.columns:
        ax.axvline(list(matrix_pct.columns).index(2023), color="blue", lw=2, ls="--", alpha=0.7)
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig03_heatmap_topics.png", bbox_inches="tight")
    plt.close()
    print("  Figure 3 : Heatmap")


def fig04_lorenz(df):
    fig, ax = plt.subplots(figsize=(10, 8))
    years_plot = [y for y in [2016, 2019, 2022, 2025] if y in df["year"].values]
    colors_l = plt.cm.viridis(np.linspace(0, 0.9, max(len(years_plot), 1)))

    for year, color in zip(years_plot, colors_l):
        all_topics = []
        for t in df[df["year"] == year]["topics_list"]:
            all_topics.extend(normalize_topics(t))
        if not all_topics:
            continue
        counts = sorted(Counter(all_topics).values())
        cum = np.cumsum(counts) / sum(counts)
        x_l = np.arange(1, len(counts) + 1) / len(counts)
        g = gini_coefficient(np.array(counts))
        ax.plot(x_l, cum, lw=2, color=color, label=str(year) + " (Gini=" + str(round(g, 3)) + ")")

    ax.plot([0, 1], [0, 1], "k--", lw=1, alpha=0.5, label="Egalite parfaite")
    ax.set_xlabel("Proportion cumulee des topics"); ax.set_ylabel("Proportion cumulee des occurrences")
    ax.set_title("Courbes de Lorenz - Concentration thematique"); ax.legend()
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig04_lorenz.png", bbox_inches="tight")
    plt.close()
    print("  Figure 4 : Lorenz")


def fig05_exposure(df):
    ai_rate = df.groupby("primary_discipline")["ai_mention"].mean()
    med = ai_rate.median()
    high = ai_rate[ai_rate >= med].index.tolist()
    low = ai_rate[ai_rate < med].index.tolist()

    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    for mi, (metric_func, ylabel) in enumerate([
        (shannon_entropy, "Shannon (bits)"),
        (lambda c: gini_coefficient(np.array(c)) if c else 0, "Gini"),
        (effective_n, "N effectif topics"),
    ]):
        ax = axes[mi]
        for gname, discs, col in [("Forte exposition IA", high, COLORS["primary"]),
                                    ("Faible exposition IA", low, COLORS["accent"])]:
            vals = []
            years = sorted(df["year"].unique())
            for y in years:
                sub = df[(df["year"] == y) & (df["primary_discipline"].isin(discs))]
                all_t = []
                for t in sub["topics_list"]:
                    all_t.extend(normalize_topics(t))
                c = list(Counter(all_t).values())
                vals.append(metric_func(c))
            ax.plot(years, vals, "o-", color=col, lw=2, ms=5, label=gname)
        ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
        ax.set_xlabel("Annee"); ax.set_ylabel(ylabel); ax.legend(fontsize=9)
        ax.set_title("(" + "abc"[mi] + ") " + ylabel)

    plt.suptitle("Diversite : disciplines forte vs faible exposition IA", fontsize=13, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig05_exposition_ia.png", bbox_inches="tight")
    plt.close()
    print("  Figure 5 : Exposition IA")
    return high, low


def fig06_before_after(metrics_q):
    mq = metrics_q.copy()
    mq["year"] = mq["period"].str[:4].astype(int)
    pre = mq[mq["year"] < 2023]
    post = mq[mq["year"] >= 2023]

    test_metrics = ["shannon", "gini", "effective_n", "hhi", "top5_share"]
    results = {}

    for m in test_metrics:
        pv = pre[m].dropna().values
        pov = post[m].dropna().values
        if len(pv) < 2 or len(pov) < 2:
            continue
        t_stat, pt = stats.ttest_ind(pv, pov, equal_var=False)
        u_stat, pu = stats.mannwhitneyu(pv, pov, alternative="two-sided")
        ps = np.sqrt((pv.std() ** 2 + pov.std() ** 2) / 2)
        cd = (pov.mean() - pv.mean()) / ps if ps > 0 else 0
        results[m] = {
            "pre_mean": pv.mean(), "pre_std": pv.std(),
            "post_mean": pov.mean(), "post_std": pov.std(),
            "diff": pov.mean() - pv.mean(),
            "diff_pct": (pov.mean() - pv.mean()) / pv.mean() * 100 if pv.mean() != 0 else 0,
            "t_stat": t_stat, "p_welch": pt, "u_stat": u_stat, "p_mwu": pu, "cohens_d": cd,
        }

    valid_metrics = [m for m in test_metrics if m in results]
    if not valid_metrics:
        print("  Figure 6 : Pas assez de donnees pour les tests")
        return results

    fig, axes = plt.subplots(1, len(valid_metrics), figsize=(4 * len(valid_metrics), 5))
    if len(valid_metrics) == 1:
        axes = [axes]
    for i, m in enumerate(valid_metrics):
        ax = axes[i]
        bp = ax.boxplot(
            [pre[m].dropna().values, post[m].dropna().values],
            labels=["Avant\n(2015-22)", "Apres\n(2023-26)"],
            patch_artist=True, medianprops=dict(color=COLORS["secondary"], lw=2)
        )
        bp["boxes"][0].set_facecolor(COLORS["light_blue"])
        bp["boxes"][1].set_facecolor(COLORS["light_red"])
        p = results[m]["p_welch"]
        sig = "***" if p < 0.01 else "**" if p < 0.05 else "*" if p < 0.1 else "ns"
        ax.set_title(m + "\n(p=" + str(round(p, 3)) + " " + sig + ")", fontsize=10)

    plt.suptitle("Comparaison avant/apres GenAI", fontsize=13, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig06_avant_apres.png", bbox_inches="tight")
    plt.close()
    print("  Figure 6 : Avant/apres")
    return results


def fig07_intensity(df):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    ax = axes[0]
    iy = df.groupby(["year", "ai_intensity"]).size().unstack(fill_value=0)
    iy_pct = iy.div(iy.sum(axis=1), axis=0) * 100
    order = [c for c in ["none", "peripheral", "methodological", "core"] if c in iy_pct.columns]
    colors_i = {"none": "#E5E7EB", "peripheral": "#93C5FD", "methodological": "#3B82F6", "core": "#1D4ED8"}
    iy_pct[order].plot.bar(stacked=True, ax=ax, color=[colors_i[c] for c in order], width=0.8)
    ax.set_xlabel("Annee"); ax.set_ylabel("% articles"); ax.set_title("(a) Intensite IA par annee")
    ax.legend(title="Intensite", fontsize=9); ax.tick_params(axis="x", rotation=45)

    ax = axes[1]
    for intensity, col in [("none", COLORS["gray"]), ("peripheral", "#93C5FD"),
                            ("methodological", "#3B82F6"), ("core", "#1D4ED8")]:
        sub = df[df["ai_intensity"] == intensity]
        if len(sub) < 10:
            continue
        ys = []
        for y in sorted(df["year"].unique()):
            all_t = []
            for t in sub[sub["year"] == y]["topics_list"]:
                all_t.extend(normalize_topics(t))
            c = list(Counter(all_t).values())
            ys.append({"year": y, "shannon": shannon_entropy(c)})
        ydf = pd.DataFrame(ys)
        ax.plot(ydf["year"], ydf["shannon"], "o-", color=col, lw=2, ms=5, label=intensity)

    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("Shannon (bits)")
    ax.set_title("(b) Diversite par intensite IA"); ax.legend(title="Intensite", fontsize=9)

    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig07_intensite_ia.png", bbox_inches="tight")
    plt.close()
    print("  Figure 7 : Intensite IA")


# =====================================================================
# PRESTIGE DES SOURCES
# =====================================================================

def compute_prestige_tiers(df):
    """Categorise les journaux par prestige (Top 10% / Top 25% / Reste)
    base sur la mediane de cited_by_count par journal.
    Modifie df en place et retourne (df, stats_df, p90, p75)."""
    journal_counts = df["source_journal"].value_counts()
    eligible = journal_counts[journal_counts >= 5].index

    journal_stats = (
        df[df["source_journal"].isin(eligible)]
        .groupby("source_journal")["cited_by_count"]
        .median()
        .reset_index()
        .rename(columns={"cited_by_count": "median_citations"})
    )

    p90 = journal_stats["median_citations"].quantile(0.90)
    p75 = journal_stats["median_citations"].quantile(0.75)

    def assign_tier(v):
        if v >= p90:
            return "Top 10%"
        if v >= p75:
            return "Top 25%"
        return "Reste"

    journal_stats["prestige_tier"] = journal_stats["median_citations"].apply(assign_tier)
    tier_map = journal_stats.set_index("source_journal")["prestige_tier"].to_dict()
    df["prestige_tier"] = df["source_journal"].map(tier_map).fillna("Reste")
    return df, journal_stats, p90, p75


def compute_prestige_metrics_by_year(df):
    """Calcule Shannon et Gini par tier de prestige et par annee."""
    results = []
    for (year, tier), group in df.groupby(["year", "prestige_tier"]):
        all_topics = []
        for topics in group["topics_list"]:
            all_topics.extend(normalize_topics(topics))
        c = list(Counter(all_topics).values())
        results.append({
            "year": int(year),
            "prestige_tier": tier,
            "n_articles": len(group),
            "pct_ai": round(group["ai_mention"].mean() * 100, 1),
            "shannon": round(shannon_entropy(c), 4) if c else 0.0,
            "gini": round(gini_coefficient(np.array(c)), 4) if c else 0.0,
            "effective_n": round(effective_n(c), 1) if c else 0.0,
        })
    return pd.DataFrame(results).sort_values(["year", "prestige_tier"]).reset_index(drop=True)


def fig08_prestige(prestige_y):
    tiers = ["Top 10%", "Top 25%", "Reste"]
    tier_colors = {
        "Top 10%": COLORS["secondary"],
        "Top 25%": COLORS["primary"],
        "Reste": COLORS["gray"],
    }
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))

    ax = axes[0]
    for tier in tiers:
        sub = prestige_y[prestige_y["prestige_tier"] == tier]
        ax.plot(sub["year"], sub["n_articles"], "o-", color=tier_colors[tier], lw=2, ms=5, label=tier)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("N articles")
    ax.set_title("(a) Volume par tier de prestige"); ax.legend(fontsize=9)

    ax = axes[1]
    for tier in tiers:
        sub = prestige_y[prestige_y["prestige_tier"] == tier]
        ax.plot(sub["year"], sub["shannon"], "o-", color=tier_colors[tier], lw=2, ms=5, label=tier)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("Shannon (bits)")
    ax.set_title("(b) Diversite thematique par prestige"); ax.legend(fontsize=9)

    ax = axes[2]
    for tier in tiers:
        sub = prestige_y[prestige_y["prestige_tier"] == tier]
        ax.plot(sub["year"], sub["gini"], "o-", color=tier_colors[tier], lw=2, ms=5, label=tier)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("Gini")
    ax.set_title("(c) Concentration thematique par prestige"); ax.legend(fontsize=9)

    plt.suptitle(
        "Standardisation thematique : revues de premier rang vs second rang (2015-2026)",
        fontsize=12, fontweight="bold", y=1.02,
    )
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig08_prestige.png", bbox_inches="tight")
    plt.close()
    print("  Figure 8 : Prestige des revues")


# =====================================================================
# PERCENTILE CITATIONS NORMALISE PAR ANNEE + SHANNON PONDEREE
# =====================================================================

def add_citation_percentile(df):
    """Calcule le percentile de citations normalise par annee (0-1).
    Evite le biais temporel : un article de 2015 tres cite ne domine
    pas artificiellement un article de 2024."""
    df = df.copy()
    def rank_within_year(group):
        group = group.copy()
        group["cit_percentile"] = group["cited_by_count"].rank(pct=True, method="average")
        return group
    df = df.groupby("year", group_keys=False).apply(rank_within_year)
    return df


def shannon_weighted(topics_iter, weights_iter):
    """Entropie de Shannon ou les probabilites p_i sont ponderees
    par le volume de citations normalise (cit_percentile) de l'article source.

    topics_iter : iterable de listes de topics (une par article)
    weights_iter : iterable de scalaires (cit_percentile de l'article)

    Si tous les poids sont nuls, retourne 0.
    """
    weighted_counts = {}
    for topics, w in zip(topics_iter, weights_iter):
        if not w or np.isnan(w):
            w = 0.0
        for t in normalize_topics(topics):
            weighted_counts[t] = weighted_counts.get(t, 0.0) + w

    total = sum(weighted_counts.values())
    if total == 0:
        return 0.0
    p = np.array(list(weighted_counts.values()), dtype=float) / total
    p = p[p > 0]
    return float(-np.sum(p * np.log2(p)))


def compute_shannon_weighted_by_year(df):
    """Retourne un DataFrame avec shannon_vol (standard) et shannon_impact (pondere)
    par annee. Permet de comparer centre de gravite volumique vs impact."""
    results = []
    for year, group in df.groupby("year"):
        # Shannon volumique standard
        all_topics_vol = []
        for t in group["topics_list"]:
            all_topics_vol.extend(normalize_topics(t))
        c_vol = list(Counter(all_topics_vol).values())
        h_vol = shannon_entropy(c_vol)

        # Shannon ponderee par cit_percentile
        h_impact = shannon_weighted(
            group["topics_list"].tolist(),
            group["cit_percentile"].tolist(),
        )

        results.append({
            "year": int(year),
            "shannon_vol": round(h_vol, 4),
            "shannon_impact": round(h_impact, 4),
            "delta_impact_vol": round(h_impact - h_vol, 4),
        })
    return pd.DataFrame(results).sort_values("year").reset_index(drop=True)


def fig_shannon_compare(shannon_cmp):
    """Figure comparative : courbe volumique vs courbe impact ponderee."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    ax = axes[0]
    ax.plot(shannon_cmp["year"], shannon_cmp["shannon_vol"], "o-",
            color=COLORS["primary"], lw=2, ms=6, label="Shannon volumique (standard)")
    ax.plot(shannon_cmp["year"], shannon_cmp["shannon_impact"], "s--",
            color=COLORS["secondary"], lw=2, ms=6, label="Shannon impact (pondere citations)")
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee")
    ax.set_ylabel("Entropie de Shannon (bits)")
    ax.set_title("(a) Diversite volumique vs diversite d'impact")
    ax.legend(fontsize=9)

    ax = axes[1]
    colors_delta = [COLORS["secondary"] if d < 0 else COLORS["accent"]
                    for d in shannon_cmp["delta_impact_vol"]]
    ax.bar(shannon_cmp["year"], shannon_cmp["delta_impact_vol"],
           color=colors_delta, edgecolor="white", lw=0.5)
    ax.axhline(0, color="black", lw=0.8)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee")
    ax.set_ylabel("Delta (impact - volumique)")
    ax.set_title("(b) Ecart : centre de gravite intellectuel vs production de masse\n"
                 "Negatif = articles cites se concentrent plus vite que la production")

    plt.suptitle(
        "Shannon ponderee : production de masse vs centre de gravite intellectuel",
        fontsize=12, fontweight="bold", y=1.02,
    )
    plt.tight_layout()
    path = FIG_DIR + "/fig10_shannon_ponderee.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    print("  Figure 10 : Shannon ponderee vs volumique")
    return path


# =====================================================================
# REGRESSION OLS - EFFET MEGA-PUBLISHER
# =====================================================================

def run_mega_publisher_regression(df, metrics_y):
    """Regression OLS pour isoler l'effet 'IA' de l'effet editeur.

    Variable dependante : shannon annuel de l'article (proxy = log(n_topics+1))
    Variables independantes :
      - year (tendance temporelle)
      - post_genai (dummy avant/apres 2023)
      - ai_mention (proxy utilisation IA)
      - is_mega_publisher (flag Mega vs Editeur Traditionnel)
      - primary_discipline (effets fixes via dummies, ref = discipline la plus frequente)

    La cardinalite de source_journal etant elevee, on cree un flag binaire
    'Mega-Publisher' pour les top-5% journaux en volume (>= p95 du nombre d'articles).
    Retourne (results_summary_str, coef_df) ou None si statsmodels absent.
    """
    try:
        import statsmodels.api as sm
    except ImportError:
        print("  AVERTISSEMENT : statsmodels absent. Installez-le via 00_setup.py.")
        return None, None

    reg_df = df.copy()

    # Variable dependante : diversite intra-article (log n_topics + 1)
    reg_df["y_diversity"] = np.log1p(reg_df["n_topics"].fillna(0))

    # Flag Mega-Publisher : journaux dans le top 5% par volume d'articles
    journal_vol = reg_df["source_journal"].value_counts()
    p95_vol = journal_vol.quantile(0.95)
    mega_set = set(journal_vol[journal_vol >= p95_vol].index)
    reg_df["is_mega_publisher"] = reg_df["source_journal"].isin(mega_set).astype(int)

    # Variables independantes
    reg_df["year_centered"] = reg_df["year"] - reg_df["year"].mean()
    reg_df["ai_mention_int"] = reg_df["ai_mention"].astype(int)

    # Effets fixes disciplines (dummies, ref = discipline la plus frequente)
    top_disc = reg_df["primary_discipline"].value_counts().idxmax()
    disc_dummies = pd.get_dummies(reg_df["primary_discipline"], prefix="disc", drop_first=False)
    if "disc_" + top_disc in disc_dummies.columns:
        disc_dummies = disc_dummies.drop(columns=["disc_" + top_disc])
    # Limiter a 15 disciplines pour eviter explosion de la matrice
    disc_cols = disc_dummies.columns.tolist()[:15]
    disc_dummies = disc_dummies[disc_cols]

    X = pd.concat([
        reg_df[["year_centered", "post_genai", "ai_mention_int", "is_mega_publisher"]],
        disc_dummies,
    ], axis=1).fillna(0)
    
    X = X.astype(float)
    y = reg_df["y_diversity"].astype(float)
    
    X = sm.add_constant(X)
    
    valid = y.notna() & X.notna().all(axis=1)
    model = sm.OLS(y[valid], X[valid]).fit(cov_type="HC3")

    coef_df = pd.DataFrame({
        "variable": model.params.index,
        "coef": model.params.values,
        "std_err": model.bse.values,
        "t_stat": model.tvalues.values,
        "p_value": model.pvalues.values,
        "conf_low": model.conf_int().iloc[:, 0].values,
        "conf_high": model.conf_int().iloc[:, 1].values,
    })
    coef_df["signif"] = coef_df["p_value"].apply(
        lambda p: "***" if p < 0.01 else "**" if p < 0.05 else "*" if p < 0.1 else "ns"
    )

    summary = (
        "R2={:.4f}  Adj-R2={:.4f}  F={:.2f}  p(F)={:.4g}  N={:,}\n"
        "Mega-Publisher : beta={:.4f}  p={:.4g}\n"
        "IA mention     : beta={:.4f}  p={:.4g}\n"
        "Post-GenAI     : beta={:.4f}  p={:.4g}"
    ).format(
        model.rsquared, model.rsquared_adj, model.fvalue, model.f_pvalue, int(valid.sum()),
        coef_df.loc[coef_df["variable"] == "is_mega_publisher", "coef"].values[0],
        coef_df.loc[coef_df["variable"] == "is_mega_publisher", "p_value"].values[0],
        coef_df.loc[coef_df["variable"] == "ai_mention_int", "coef"].values[0],
        coef_df.loc[coef_df["variable"] == "ai_mention_int", "p_value"].values[0],
        coef_df.loc[coef_df["variable"] == "post_genai", "coef"].values[0],
        coef_df.loc[coef_df["variable"] == "post_genai", "p_value"].values[0],
    )
    print("  OLS mega-publisher :")
    for line in summary.split("\n"):
        print("    " + line)

    return summary, coef_df


def fig_ols_coefs(coef_df):
    """Forest plot des coefficients OLS (hors dummies disciplines et constante)."""
    if coef_df is None:
        return
    sub = coef_df[~coef_df["variable"].str.startswith("disc_") &
                  (coef_df["variable"] != "const")].copy()
    sub = sub.sort_values("coef")
    colors_c = [COLORS["secondary"] if p < 0.05 else COLORS["gray"] for p in sub["p_value"]]

    fig, ax = plt.subplots(figsize=(10, max(4, len(sub) * 0.5)))
    y_pos = range(len(sub))
    ax.barh(y_pos, sub["coef"], xerr=[sub["coef"] - sub["conf_low"],
                                       sub["conf_high"] - sub["coef"]],
            color=colors_c, height=0.5, capsize=3)
    ax.axvline(0, color="black", lw=0.8, ls="--")
    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(sub["variable"].tolist())
    ax.set_xlabel("Coefficient OLS (log n_topics)")
    ax.set_title("Effet Mega-Publisher vs effet IA\n"
                 "(rouge = significatif p<0.05, IC 95% HC3-robuste)")
    plt.tight_layout()
    path = FIG_DIR + "/fig11_ols_mega_publisher.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    print("  Figure 11 : Forest plot OLS")


# =====================================================================
# BIAIS GEOGRAPHIQUE
# =====================================================================

def compute_geo_metrics_by_year(df):
    """Calcule taux d'adoption IA et Shannon par geo_zone et par annee."""
    results = []
    for (year, zone), group in df.groupby(["year", "geo_zone"]):
        all_topics = []
        for topics in group["topics_list"]:
            all_topics.extend(normalize_topics(topics))
        c = list(Counter(all_topics).values())
        results.append({
            "year": int(year),
            "geo_zone": zone,
            "n_articles": len(group),
            "pct_ai": round(group["ai_mention"].mean() * 100, 1),
            "shannon": round(shannon_entropy(c), 4) if c else 0.0,
            "gini": round(gini_coefficient(np.array(c)), 4) if c else 0.0,
        })
    return pd.DataFrame(results).sort_values(["year", "geo_zone"]).reset_index(drop=True)


# =====================================================================
# RESEAU DE COLLABORATION INTERNATIONALE
# =====================================================================

def compute_country_cooccurrence(df, top_n=20):
    """Calcule la matrice de co-occurrence des pays et identifie les paires
    qui collaborent le plus frequemment.

    Retourne (matrix_df, country_counts, top_pairs_df).
    """
    cooc = {}
    country_counts = Counter()

    for _, row in df.iterrows():
        countries = list(set(parse_countries_for_cooc(row["countries"])))
        for c in countries:
            country_counts[c] += 1
        for c1, c2 in combinations(sorted(countries), 2):
            cooc[(c1, c2)] = cooc.get((c1, c2), 0) + 1

    top_countries = [c for c, _ in country_counts.most_common(top_n)]
    matrix = pd.DataFrame(0, index=top_countries, columns=top_countries)
    for (c1, c2), count in cooc.items():
        if c1 in matrix.index and c2 in matrix.columns:
            matrix.loc[c1, c2] = count
            matrix.loc[c2, c1] = count

    top_pairs = sorted(
        [(c1, c2, cnt) for (c1, c2), cnt in cooc.items()],
        key=lambda x: -x[2],
    )[:50]
    top_pairs_df = pd.DataFrame(top_pairs, columns=["pays_1", "pays_2", "co_occurrences"])

    return matrix, country_counts, top_pairs_df


# =====================================================================
# DYNAMIQUE DES EQUIPES DE RECHERCHE
# =====================================================================

def compute_team_dynamics(df):
    """Analyse temporelle de la taille des equipes et l'impact selon le type
    de collaboration (national vs international).

    Requiert cit_percentile dans df (calcule par add_citation_percentile).
    Retourne (team_time, collab_impact, team_by_year).
    """
    df2 = df.copy()
    df2["n_collab_countries"] = df2["countries"].apply(
        lambda v: len(set(parse_countries_for_cooc(v)))
    )
    df2["collab_type"] = df2["n_collab_countries"].apply(
        lambda n: "International" if n > 1 else "National"
    )
    df2["team_size_cat"] = pd.cut(
        df2["n_authors"].fillna(1).clip(lower=1).astype(float),
        bins=[0, 1, 3, 6, 10, float("inf")],
        labels=["Solo", "2-3", "4-6", "7-10", "11+"],
    ).astype(str).replace("nan", "Unknown")

    team_time = (
        df2.groupby("year")
        .agg(
            n_articles=("openalex_id", "count"),
            mean_authors=("n_authors", "mean"),
            median_authors=("n_authors", "median"),
        )
        .reset_index()
    )
    team_time["mean_authors"] = team_time["mean_authors"].round(2)
    team_time["median_authors"] = team_time["median_authors"].round(1)

    collab_impact = (
        df2.groupby(["collab_type", "team_size_cat"])
        .agg(
            n=("openalex_id", "count"),
            mean_cit_pct=("cit_percentile", "mean"),
            median_cit_pct=("cit_percentile", "median"),
        )
        .reset_index()
    )
    collab_impact["mean_cit_pct"] = collab_impact["mean_cit_pct"].round(4)
    collab_impact["median_cit_pct"] = collab_impact["median_cit_pct"].round(4)

    team_by_year = (
        df2.groupby(["year", "collab_type"])
        .agg(
            n=("openalex_id", "count"),
            mean_authors=("n_authors", "mean"),
            median_cit_pct=("cit_percentile", "median"),
        )
        .reset_index()
    )
    team_by_year["mean_authors"] = team_by_year["mean_authors"].round(2)
    team_by_year["median_cit_pct"] = team_by_year["median_cit_pct"].round(4)

    return team_time, collab_impact, team_by_year


# =====================================================================
# CO-OCCURRENCE DES CONCEPTS
# =====================================================================

def compute_concept_cooccurrence(df, top_n=30):
    """Identifie les intersections disciplinaires emergentes via la co-occurrence
    des concepts OpenAlex (colonne concepts_json).

    Retourne (top_concepts_df, top_pairs_df).
    """
    if "concepts_json" not in df.columns:
        print("  AVERTISSEMENT : colonne concepts_json absente.")
        empty_c = pd.DataFrame(columns=["concept", "n_articles"])
        empty_p = pd.DataFrame(columns=["concept_1", "concept_2", "co_occurrences"])
        return empty_c, empty_p

    cooc = {}
    concept_counts = Counter()

    for _, row in df.iterrows():
        concepts = list(set(parse_concepts(row.get("concepts_json"))))
        for c in concepts:
            concept_counts[c] += 1
        for c1, c2 in combinations(sorted(concepts), 2):
            cooc[(c1, c2)] = cooc.get((c1, c2), 0) + 1

    top_concepts_df = pd.DataFrame(
        concept_counts.most_common(top_n * 2),
        columns=["concept", "n_articles"],
    )
    top_pairs = sorted(
        [(c1, c2, cnt) for (c1, c2), cnt in cooc.items()],
        key=lambda x: -x[2],
    )[: top_n * 3]
    top_pairs_df = pd.DataFrame(top_pairs, columns=["concept_1", "concept_2", "co_occurrences"])

    return top_concepts_df, top_pairs_df


def fig09_geographie(geo_y):
    zones = ["Global North", "Global South", "International"]
    zone_colors = {
        "Global North": COLORS["primary"],
        "Global South": COLORS["secondary"],
        "International": COLORS["accent"],
        "Unknown": COLORS["gray"],
    }
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    ax = axes[0]
    for zone in zones:
        sub = geo_y[geo_y["geo_zone"] == zone]
        if sub.empty:
            continue
        ax.plot(sub["year"], sub["pct_ai"], "o-", color=zone_colors[zone], lw=2, ms=5, label=zone)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("% mention IA")
    ax.set_title("(a) Taux d'adoption IA par zone geographique"); ax.legend(fontsize=9)

    ax = axes[1]
    for zone in zones:
        sub = geo_y[geo_y["geo_zone"] == zone]
        if sub.empty:
            continue
        ax.plot(sub["year"], sub["shannon"], "o-", color=zone_colors[zone], lw=2, ms=5, label=zone)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee"); ax.set_ylabel("Shannon (bits)")
    ax.set_title("(b) Diversite thematique par zone geographique"); ax.legend(fontsize=9)

    plt.suptitle(
        "Biais geographique : adoption IA et diversite taxonomique (2015-2026)",
        fontsize=12, fontweight="bold", y=1.02,
    )
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig09_geographie.png", bbox_inches="tight")
    plt.close()
    print("  Figure 9 : Biais geographique")


def fig12_countries_network(cooc_matrix):
    """Heatmap des 20 pays les plus interconnectes (co-publications)."""
    if cooc_matrix.empty:
        print("  Figure 12 : Pas de donnees pays")
        return

    log_matrix = np.log1p(cooc_matrix.astype(float))
    mask = cooc_matrix == 0

    fig, ax = plt.subplots(figsize=(14, 12))
    sns.heatmap(
        log_matrix,
        mask=mask,
        cmap="Blues",
        ax=ax,
        linewidths=0.4,
        linecolor="#E5E7EB",
        cbar_kws={"label": "Co-occurrences (log1p)", "shrink": 0.8},
        annot=False,
    )
    ax.set_title(
        "Reseau de collaboration internationale\n(top 20 pays par volume, echelle log)",
        fontsize=13,
        fontweight="bold",
    )
    ax.tick_params(axis="x", rotation=45, labelsize=9)
    ax.tick_params(axis="y", rotation=0, labelsize=9)
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig12_countries_network.png", bbox_inches="tight")
    plt.close()
    print("  Figure 12 : Reseau de collaboration internationale")


def fig13_team_dynamics(team_time, team_by_year, collab_impact):
    """Figure composite : evolution taille des equipes et prime a la collaboration."""
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))

    # (a) Evolution temporelle de la taille des equipes
    ax = axes[0]
    ax.plot(team_time["year"], team_time["mean_authors"], "o-",
            color=COLORS["primary"], lw=2, ms=5, label="Moyenne")
    ax.plot(team_time["year"], team_time["median_authors"], "s--",
            color=COLORS["secondary"], lw=2, ms=5, label="Mediane")
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee")
    ax.set_ylabel("N auteurs")
    ax.set_title("(a) Evolution de la taille des equipes")
    ax.legend(fontsize=9)

    # (b) Impact (cit_percentile median) par type de collaboration, par annee
    ax = axes[1]
    for collab_type, color in [("International", COLORS["primary"]),
                                ("National", COLORS["accent"])]:
        sub = team_by_year[team_by_year["collab_type"] == collab_type].sort_values("year")
        if sub.empty:
            continue
        ax.plot(sub["year"], sub["median_cit_pct"], "o-",
                color=color, lw=2, ms=5, label=collab_type)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee")
    ax.set_ylabel("Percentile citations (mediane)")
    ax.set_title("(b) Impact selon le type de collaboration")
    ax.legend(fontsize=9)

    # (c) Prime a la collaboration : impact median par taille d'equipe
    ax = axes[2]
    team_sizes = ["Solo", "2-3", "4-6", "7-10", "11+"]
    x = np.arange(len(team_sizes))
    width = 0.35
    for i, (ctype, color) in enumerate([("National", COLORS["accent"]),
                                         ("International", COLORS["primary"])]):
        vals = []
        for ts in team_sizes:
            sub = collab_impact[
                (collab_impact["collab_type"] == ctype) &
                (collab_impact["team_size_cat"] == ts)
            ]
            vals.append(float(sub["median_cit_pct"].iloc[0]) if not sub.empty else 0.0)
        ax.bar(x + i * width, vals, width, label=ctype, color=color,
               alpha=0.85, edgecolor="white")
    ax.set_xticks(x + width / 2)
    ax.set_xticklabels(team_sizes, fontsize=9)
    ax.set_xlabel("Taille d'equipe (N auteurs)")
    ax.set_ylabel("Percentile citations (mediane)")
    ax.set_title("(c) Prime a la collaboration\npar taille d'equipe")
    ax.legend(fontsize=9)

    plt.suptitle(
        "Dynamique des equipes : taille, impact et type de collaboration (2015-2026)",
        fontsize=12, fontweight="bold", y=1.02,
    )
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig13_team_dynamics.png", bbox_inches="tight")
    plt.close()
    print("  Figure 13 : Dynamique des equipes de recherche")


# =====================================================================
# EXCEL FINAL
# =====================================================================

def create_analysis_excel(df, metrics_q, metrics_y, test_results, high_ai, low_ai,
                           prestige_y, journal_stats, geo_y, shannon_cmp, ols_summary,
                           coef_df, country_cooc_matrix, country_top_pairs,
                           team_time, team_by_year, collab_impact,
                           concept_top_df, concept_pairs_df, output_path):
    wb = Workbook()

    # --- 1. Stats descriptives ---
    ws = wb.active
    ws.title = "1. Stats descriptives"
    ws.sheet_properties.tabColor = "2563EB"
    write_title(ws, 1, "STATISTIQUES DESCRIPTIVES GLOBALES")

    stat_headers = ["Metrique", "2015-2022", "2023-2026", "Total"]
    write_header_row(ws, 3, stat_headers)

    stat_labels = ["N articles", "N disciplines", "N topics uniques", "Topics/article (moy)",
                    "Citations (mediane)", "% mention IA", "Shannon entropy", "Simpson diversity",
                    "Gini", "HHI", "Top-5 share (%)", "Top-10 share (%)"]

    stat_data = []
    for label in stat_labels:
        row_vals = [label]
        for mask in [df["year"] < 2023, df["year"] >= 2023, df["year"] > 0]:
            sub = df[mask]
            all_t = []
            for t in sub["topics_list"]:
                all_t.extend(normalize_topics(t))
            c = list(Counter(all_t).values())

            val_map = {
                "N articles": len(sub),
                "N disciplines": sub["primary_discipline"].nunique(),
                "N topics uniques": len(Counter(all_t)),
                "Topics/article (moy)": round(sub["n_topics"].mean(), 2) if "n_topics" in sub.columns else "N/A",
                "Citations (mediane)": round(sub["cited_by_count"].median(), 1),
                "% mention IA": round(sub["ai_mention"].mean() * 100, 1),
                "Shannon entropy": round(shannon_entropy(c), 4),
                "Simpson diversity": round(simpson_diversity(c), 4),
                "Gini": round(gini_coefficient(np.array(c)), 4),
                "HHI": round(hhi_index(c), 6),
                "Top-5 share (%)": round(top_n_share(c, 5) * 100, 2),
                "Top-10 share (%)": round(top_n_share(c, 10) * 100, 2),
            }
            row_vals.append(val_map[label])
        stat_data.append(row_vals)

    write_data_rows(ws, 4, stat_data, 4)
    set_col_widths(ws, {"A": 25, "B": 15, "C": 15, "D": 15})

    # --- 2. Volume par annee ---
    ws2 = wb.create_sheet("2. Volume par annee")
    ws2.sheet_properties.tabColor = "059669"
    write_title(ws2, 1, "VOLUME D'ARTICLES PAR ANNEE")
    yearly = df.groupby("year").agg(n=("openalex_id", "count"), n_ai=("ai_mention", "sum"),
                                     cit=("cited_by_count", "median"), auth=("n_authors", "mean")).reset_index()
    yearly["pct"] = (yearly["n_ai"] / yearly["n"] * 100).round(1)
    h2 = ["Annee", "Total", "IA", "% IA", "Citations med.", "Auteurs moy."]
    write_header_row(ws2, 3, h2, "059669")
    d2 = [[int(r["year"]), int(r["n"]), int(r["n_ai"]), r["pct"], round(r["cit"], 1), round(r["auth"], 1)] for _, r in yearly.iterrows()]
    write_data_rows(ws2, 4, d2, 6)
    set_col_widths(ws2, {"A": 10, "B": 10, "C": 10, "D": 10, "E": 16, "F": 16})

    # --- 3. Diversite trimestrielle ---
    ws3 = wb.create_sheet("3. Diversite trim.")
    ws3.sheet_properties.tabColor = "7C3AED"
    write_title(ws3, 1, "METRIQUES DE DIVERSITE PAR TRIMESTRE")
    h3 = list(metrics_q.columns)
    write_header_row(ws3, 3, h3, "7C3AED")
    d3 = [list(row) for _, row in metrics_q.iterrows()]
    write_data_rows(ws3, 4, d3, len(h3))
    for i, c in enumerate(h3):
        ws3.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)].width = max(14, len(str(c)) + 3)

    # --- 4. Diversite annuelle ---
    ws4 = wb.create_sheet("4. Diversite annuelle")
    ws4.sheet_properties.tabColor = "DC2626"
    write_title(ws4, 1, "METRIQUES DE DIVERSITE PAR ANNEE")
    h4 = list(metrics_y.columns)
    write_header_row(ws4, 3, h4, "DC2626")
    d4 = [list(row) for _, row in metrics_y.iterrows()]
    write_data_rows(ws4, 4, d4, len(h4))
    for i, c in enumerate(h4):
        ws4.column_dimensions[chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)].width = max(14, len(str(c)) + 3)

    # --- 5. Concentration ---
    ws5 = wb.create_sheet("5. Concentration")
    ws5.sheet_properties.tabColor = "F59E0B"
    write_title(ws5, 1, "INDICATEURS DE CONCENTRATION PAR ANNEE")
    h5 = ["Annee", "Gini", "HHI", "Top-5 (%)", "Top-10 (%)", "N topics"]
    write_header_row(ws5, 3, h5, "F59E0B")
    d5 = [[r["period"], r["gini"], r["hhi"], r["top5_share"], r["top10_share"], r["n_unique_topics"]] for _, r in metrics_y.iterrows()]
    write_data_rows(ws5, 4, d5, 6)
    set_col_widths(ws5, {"A": 10, "B": 10, "C": 12, "D": 12, "E": 12, "F": 12})

    # --- 6. Tests avant/apres ---
    ws6 = wb.create_sheet("6. Tests avant-apres")
    ws6.sheet_properties.tabColor = "EC4899"
    write_title(ws6, 1, "TESTS STATISTIQUES AVANT vs APRES GenAI (2023)")
    h6 = ["Metrique", "Moy avant", "ET avant", "Moy apres", "ET apres", "Delta", "Delta %", "t Welch", "p-value", "Cohen d", "Signif."]
    write_header_row(ws6, 3, h6, "EC4899")
    d6 = []
    for m, r in test_results.items():
        sig = "***" if r["p_welch"] < 0.01 else "**" if r["p_welch"] < 0.05 else "*" if r["p_welch"] < 0.1 else "ns"
        d6.append([m, round(r["pre_mean"], 4), round(r["pre_std"], 4), round(r["post_mean"], 4),
                    round(r["post_std"], 4), round(r["diff"], 4), round(r["diff_pct"], 1),
                    round(r["t_stat"], 3), round(r["p_welch"], 4), round(r["cohens_d"], 3), sig])
    write_data_rows(ws6, 4, d6, 11)
    for i in range(11):
        ws6.column_dimensions[chr(65 + i)].width = 14

    # --- 7. Exposition IA ---
    ws7 = wb.create_sheet("7. Exposition IA")
    ws7.sheet_properties.tabColor = "2563EB"
    write_title(ws7, 1, "DIVERSITE PAR NIVEAU D'EXPOSITION IA")
    ws7.cell(row=2, column=1, value="Forte : " + ", ".join(high_ai[:6])).font = Font(name="Arial", size=9, color="6B7280")
    ws7.cell(row=3, column=1, value="Faible : " + ", ".join(low_ai[:6])).font = Font(name="Arial", size=9, color="6B7280")

    h7 = ["Annee", "Shannon forte", "Shannon faible", "Gini forte", "Gini faible", "N eff. forte", "N eff. faible"]
    write_header_row(ws7, 5, h7)
    d7 = []
    for y in sorted(df["year"].unique()):
        rv = [int(y)]
        for metric_func in [shannon_entropy, lambda c: gini_coefficient(np.array(c)) if c else 0, effective_n]:
            for discs in [high_ai, low_ai]:
                sub = df[(df["year"] == y) & (df["primary_discipline"].isin(discs))]
                all_t = []
                for t in sub["topics_list"]:
                    all_t.extend(normalize_topics(t))
                c = list(Counter(all_t).values())
                rv.append(round(metric_func(c), 4))
        d7.append(rv)
    write_data_rows(ws7, 6, d7, 7)
    set_col_widths(ws7, {"A": 10, "B": 16, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14})

    # --- 8. Intensite IA ---
    ws8 = wb.create_sheet("8. Intensite IA")
    ws8.sheet_properties.tabColor = "059669"
    write_title(ws8, 1, "DIVERSITE PAR INTENSITE DE MENTION IA")
    h8 = ["Annee", "Shannon none", "Shannon peripheral", "Shannon method.", "Shannon core"]
    write_header_row(ws8, 3, h8, "059669")
    d8 = []
    for y in sorted(df["year"].unique()):
        rv = [int(y)]
        for lvl in ["none", "peripheral", "methodological", "core"]:
            sub = df[(df["year"] == y) & (df["ai_intensity"] == lvl)]
            all_t = []
            for t in sub["topics_list"]:
                all_t.extend(normalize_topics(t))
            c = list(Counter(all_t).values())
            rv.append(round(shannon_entropy(c), 4) if c else "N/A")
        d8.append(rv)
    write_data_rows(ws8, 4, d8, 5)
    set_col_widths(ws8, {"A": 10, "B": 18, "C": 20, "D": 20, "E": 16})

    # --- 9. Top 30 topics ---
    ws9 = wb.create_sheet("9. Top 30 topics")
    ws9.sheet_properties.tabColor = "7C3AED"
    write_title(ws9, 1, "TOP 30 TOPICS - FREQUENCE ET EVOLUTION")
    all_topics_global = []
    for t in df["topics_list"]:
        all_topics_global.extend(normalize_topics(t))
    top30 = Counter(all_topics_global).most_common(30)

    # Pre-count topics by period for efficiency
    pre_topics = []
    post_topics = []
    for _, r in df.iterrows():
        tl = normalize_topics(r["topics_list"])
        if r["year"] < 2023:
            pre_topics.extend(tl)
        else:
            post_topics.extend(tl)
    pre_counter = Counter(pre_topics)
    post_counter = Counter(post_topics)

    h9 = ["Rang", "Topic", "Total", "Avant 2023", "Apres 2023"]
    write_header_row(ws9, 3, h9, "7C3AED")
    d9 = []
    for rank, (topic, total) in enumerate(top30, 1):
        d9.append([rank, topic, total, pre_counter.get(topic, 0), post_counter.get(topic, 0)])
    write_data_rows(ws9, 4, d9, 5)
    set_col_widths(ws9, {"A": 8, "B": 35, "C": 10, "D": 14, "E": 14})

    # --- 10. Methodologie ---
    ws10 = wb.create_sheet("10. Methodologie")
    ws10.sheet_properties.tabColor = "6B7280"
    write_title(ws10, 1, "NOTES METHODOLOGIQUES")

    notes = [
        ["Source", "OpenAlex API (https://openalex.org) - base ouverte, ~250M works"],
        ["Periode", "2015-2026 (12 annees, avant/apres IA generative)"],
        ["Echantillonnage", "Aleatoire reproductible (seed=42) parmi articles avec abstract"],
        ["Proxy IA", "Detection par regex de 45+ mots-cles dans les abstracts"],
        ["Intensite", "none (0 match) / peripheral (1-2) / methodological (3-5) / core (6+)"],
        ["Shannon", "H = -Sum p_i log2(p_i) - diversite des topics"],
        ["Simpson", "1 - Sum p_i^2 - proba que 2 articles aleatoires different"],
        ["N effectif", "2^H - nombre de topics equivalents si distribution uniforme"],
        ["Gini", "Inegalite de distribution. 0=uniforme, 1=un seul topic domine"],
        ["HHI", "Herfindahl-Hirschman = Sum s_i^2 - concentration (standard IO)"],
        ["Top-N share", "Part cumulee des N topics les plus frequents"],
        ["Lorenz", "Courbe cumulee ordonnee, comparee a ligne d'egalite"],
        ["Welch t-test", "Test bilateral robuste (variances inegales) avant vs apres"],
        ["Cohen d", "Taille d'effet. |d|>0.8=grand, 0.5=moyen, 0.2=petit"],
        ["Exposition", "Discipline forte si taux mention IA >= mediane inter-disciplines"],
    ]

    h10 = ["Concept", "Description"]
    write_header_row(ws10, 3, h10, "6B7280")
    write_data_rows(ws10, 4, notes, 2)
    # Make description cells wrap
    for i in range(len(notes)):
        ws10.cell(row=4 + i, column=1).font = Font(name="Arial", size=10, bold=True)
        ws10.cell(row=4 + i, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    set_col_widths(ws10, {"A": 20, "B": 80})

    # --- 11. Prestige des revues ---
    ws11 = wb.create_sheet("11. Prestige revues")
    ws11.sheet_properties.tabColor = "DC2626"
    write_title(ws11, 1, "DIVERSITE THEMATIQUE PAR PRESTIGE DE REVUE")
    ws11.cell(row=2, column=1, value="Top 10% / Top 25% bases sur la mediane cited_by_count par journal (min. 5 articles)").font = Font(name="Arial", size=9, italic=True, color="6B7280")
    h11 = ["Annee", "Tier prestige", "N articles", "% IA", "Shannon", "Gini", "N effectif"]
    write_header_row(ws11, 4, h11, "DC2626")
    d11 = [
        [r["year"], r["prestige_tier"], r["n_articles"], r["pct_ai"],
         r["shannon"], r["gini"], r["effective_n"]]
        for _, r in prestige_y.iterrows()
    ]
    write_data_rows(ws11, 5, d11, 7)
    set_col_widths(ws11, {"A": 10, "B": 14, "C": 12, "D": 10, "E": 12, "F": 10, "G": 12})

    # Top journaux par tier (info complementaire)
    if not journal_stats.empty:
        note_row = 5 + len(d11) + 2
        ws11.cell(row=note_row, column=1, value="Top 5 journaux par tier :").font = Font(name="Arial", bold=True, size=10)
        nr = note_row + 1
        for tier in ["Top 10%", "Top 25%", "Reste"]:
            top5 = journal_stats[journal_stats["prestige_tier"] == tier].nlargest(5, "median_citations")
            ws11.cell(row=nr, column=1, value=tier).font = Font(name="Arial", bold=True, size=9, color="6B7280")
            for j, (_, jr) in enumerate(top5.iterrows()):
                ws11.cell(row=nr, column=2 + j, value=str(jr["source_journal"])[:40]).font = Font(name="Arial", size=9)
            nr += 1

    # --- 12. Biais geographique ---
    ws12 = wb.create_sheet("12. Biais geographique")
    ws12.sheet_properties.tabColor = "059669"
    write_title(ws12, 1, "ADOPTION IA ET DIVERSITE PAR ZONE GEOGRAPHIQUE")
    ws12.cell(row=2, column=1, value="Global North = OCDE + hauts revenus | Global South = reste | International = co-publication Nord-Sud").font = Font(name="Arial", size=9, italic=True, color="6B7280")
    h12 = ["Annee", "Zone geographique", "N articles", "% IA", "Shannon", "Gini"]
    write_header_row(ws12, 4, h12, "059669")
    d12 = [
        [r["year"], r["geo_zone"], r["n_articles"], r["pct_ai"], r["shannon"], r["gini"]]
        for _, r in geo_y.iterrows()
    ]
    write_data_rows(ws12, 5, d12, 6)
    set_col_widths(ws12, {"A": 10, "B": 18, "C": 12, "D": 10, "E": 12, "F": 10})

    # --- 13. Shannon ponderee ---
    ws13 = wb.create_sheet("13. Shannon ponderee")
    ws13.sheet_properties.tabColor = "7C3AED"
    write_title(ws13, 1, "SHANNON PONDEREE PAR CITATIONS : VOLUMIQUE vs IMPACT")
    ws13.cell(row=2, column=1,
              value="shannon_vol = entropie standard | shannon_impact = ponderee par cit_percentile annuel").font = Font(name="Arial", size=9, italic=True, color="6B7280")
    h13 = ["Annee", "Shannon volumique", "Shannon impact", "Delta (impact - vol)",
           "Interpretation"]
    write_header_row(ws13, 4, h13, "7C3AED")
    d13 = []
    for _, r in shannon_cmp.iterrows():
        delta = r["delta_impact_vol"]
        if delta < -0.05:
            interp = "Articles cites plus concentres que la production"
        elif delta > 0.05:
            interp = "Articles cites plus diversifies que la production"
        else:
            interp = "Centre de gravite aligne sur la production"
        d13.append([int(r["year"]), r["shannon_vol"], r["shannon_impact"],
                    round(delta, 4), interp])
    write_data_rows(ws13, 5, d13, 5)
    set_col_widths(ws13, {"A": 10, "B": 20, "C": 20, "D": 22, "E": 50})

    # --- 14. Regression OLS ---
    ws14 = wb.create_sheet("14. Regression OLS")
    ws14.sheet_properties.tabColor = "EC4899"
    write_title(ws14, 1, "REGRESSION OLS : EFFET IA vs EFFET MEGA-PUBLISHER")
    ws14.cell(row=2, column=1,
              value="VD : log(n_topics+1) | VI : year, post_genai, ai_mention, is_mega_publisher, effets fixes disciplines").font = Font(name="Arial", size=9, italic=True, color="6B7280")
    if ols_summary:
        for i, line in enumerate(ols_summary.split("\n")):
            ws14.cell(row=3 + i, column=1, value=line).font = Font(name="Arial", size=10, italic=True, color="2563EB")
    if coef_df is not None:
        h14 = ["Variable", "Coefficient", "Std Err", "t-stat", "p-value",
               "IC bas (95%)", "IC haut (95%)", "Significativite"]
        write_header_row(ws14, 8, h14, "EC4899")
        d14 = []
        for _, r in coef_df.iterrows():
            d14.append([
                str(r["variable"]),
                round(r["coef"], 6), round(r["std_err"], 6),
                round(r["t_stat"], 3), round(r["p_value"], 4),
                round(r["conf_low"], 6), round(r["conf_high"], 6),
                r["signif"],
            ])
        write_data_rows(ws14, 9, d14, 8)
        set_col_widths(ws14, {"A": 30, "B": 14, "C": 12, "D": 10,
                               "E": 12, "F": 14, "G": 14, "H": 16})
    else:
        ws14.cell(row=9, column=1,
                  value="statsmodels absent - installez via : pip install statsmodels").font = Font(name="Arial", size=11, color="DC2626")

    # --- 15. Reseau de collaboration internationale ---
    ws15 = wb.create_sheet("15. Collab internationale")
    ws15.sheet_properties.tabColor = "2563EB"
    write_title(ws15, 1, "RESEAU DE COLLABORATION INTERNATIONALE (TOP 20 PAYS)")
    ws15.cell(row=2, column=1,
              value="Matrice de co-occurrence : nb d'articles co-publies entre paires de pays (echelle brute)").font = Font(name="Arial", size=9, italic=True, color="6B7280")

    if not country_cooc_matrix.empty:
        countries_list = country_cooc_matrix.index.tolist()
        # En-tetes colonnes
        ws15.cell(row=4, column=1, value="Pays").font = HEADER_FONT
        ws15.cell(row=4, column=1).fill = PatternFill("solid", fgColor="2563EB")
        ws15.cell(row=4, column=1).border = THIN_BORDER
        for j, c in enumerate(countries_list, 2):
            cell = ws15.cell(row=4, column=j, value=c)
            cell.font = HEADER_FONT
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        # Lignes de donnees
        for i, c1 in enumerate(countries_list):
            lbl = ws15.cell(row=5 + i, column=1, value=c1)
            lbl.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            lbl.fill = PatternFill("solid", fgColor="2563EB")
            lbl.border = THIN_BORDER
            for j, c2 in enumerate(countries_list):
                v = int(country_cooc_matrix.loc[c1, c2])
                cell = ws15.cell(row=5 + i, column=2 + j, value=v if v > 0 else "")
                cell.font = DATA_FONT
                cell.alignment = DATA_ALIGN
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_FILL
        for k in range(len(countries_list) + 1):
            ws15.column_dimensions[get_column_letter(k + 1)].width = 8

    # Top 50 paires
    pairs_start = 5 + len(country_cooc_matrix) + 3 if not country_cooc_matrix.empty else 4
    ws15.cell(row=pairs_start, column=1,
              value="Top 50 paires de pays (co-publications)").font = Font(name="Arial", bold=True, size=11)
    if not country_top_pairs.empty:
        h15b = ["Pays 1", "Pays 2", "Co-occurrences"]
        write_header_row(ws15, pairs_start + 1, h15b, "2563EB")
        d15b = [
            [r["pays_1"], r["pays_2"], int(r["co_occurrences"])]
            for _, r in country_top_pairs.iterrows()
        ]
        write_data_rows(ws15, pairs_start + 2, d15b, 3)
        ws15.column_dimensions["A"].width = 12
        ws15.column_dimensions["B"].width = 12
        ws15.column_dimensions["C"].width = 18

    # --- 16. Equipes de recherche ---
    ws16 = wb.create_sheet("16. Equipes recherche")
    ws16.sheet_properties.tabColor = "059669"
    write_title(ws16, 1, "DYNAMIQUE DES EQUIPES DE RECHERCHE : TAILLE ET IMPACT")
    ws16.cell(row=2, column=1,
              value="International = >1 pays | National = 1 pays | cit_pct = percentile citations normalise par annee").font = Font(name="Arial", size=9, italic=True, color="6B7280")

    # Section A : evolution temporelle de la taille des equipes
    ws16.cell(row=4, column=1,
              value="A. Evolution temporelle de la taille des equipes").font = Font(name="Arial", bold=True, size=11)
    h16a = ["Annee", "N articles", "Auteurs (moy.)", "Auteurs (med.)"]
    write_header_row(ws16, 5, h16a, "059669")
    d16a = [
        [int(r["year"]), int(r["n_articles"]), r["mean_authors"], r["median_authors"]]
        for _, r in team_time.iterrows()
    ]
    write_data_rows(ws16, 6, d16a, 4)
    set_col_widths(ws16, {"A": 10, "B": 12, "C": 16, "D": 16})

    # Section B : impact par type de collaboration et annee
    b_start = 6 + len(d16a) + 3
    ws16.cell(row=b_start, column=1,
              value="B. Impact (percentile citations median) par type de collaboration").font = Font(name="Arial", bold=True, size=11)
    h16b = ["Annee", "Type collab.", "N articles", "Auteurs (moy.)", "Percentile cit. med."]
    write_header_row(ws16, b_start + 1, h16b, "059669")
    d16b = [
        [int(r["year"]), r["collab_type"], int(r["n"]),
         r["mean_authors"], r["median_cit_pct"]]
        for _, r in team_by_year.sort_values(["year", "collab_type"]).iterrows()
    ]
    write_data_rows(ws16, b_start + 2, d16b, 5)

    # Section C : prime a la collaboration
    c_start = b_start + 2 + len(d16b) + 3
    ws16.cell(row=c_start, column=1,
              value="C. Prime a la collaboration : impact median par taille d'equipe").font = Font(name="Arial", bold=True, size=11)
    h16c = ["Type collab.", "Taille equipe", "N articles", "Pct cit. moyen", "Pct cit. median"]
    write_header_row(ws16, c_start + 1, h16c, "059669")
    cat_order = ["Solo", "2-3", "4-6", "7-10", "11+", "Unknown"]
    d16c = []
    for _, r in collab_impact.iterrows():
        d16c.append([
            r["collab_type"], r["team_size_cat"], int(r["n"]),
            r["mean_cit_pct"], r["median_cit_pct"],
        ])
    # Tri par type puis taille
    d16c.sort(key=lambda x: (x[0], cat_order.index(x[1]) if x[1] in cat_order else 99))
    write_data_rows(ws16, c_start + 2, d16c, 5)
    for col in ["A", "B", "C", "D", "E"]:
        ws16.column_dimensions[col].width = 18

    # --- 17. Co-occurrence des concepts ---
    ws17 = wb.create_sheet("17. Concepts co-occurrence")
    ws17.sheet_properties.tabColor = "7C3AED"
    write_title(ws17, 1, "MULTIDISCIPLINARITE ET CO-OCCURRENCE DES CONCEPTS")
    ws17.cell(row=2, column=1,
              value="Source : colonne concepts_json (OpenAlex) | Intersections disciplinaires emergentes").font = Font(name="Arial", size=9, italic=True, color="6B7280")

    # Section A : top concepts par frequence
    ws17.cell(row=4, column=1,
              value="A. Top concepts par frequence d'apparition").font = Font(name="Arial", bold=True, size=11)
    h17a = ["Rang", "Concept", "N articles"]
    write_header_row(ws17, 5, h17a, "7C3AED")
    if not concept_top_df.empty:
        d17a = [
            [rank + 1, r["concept"], int(r["n_articles"])]
            for rank, (_, r) in enumerate(concept_top_df.iterrows())
        ]
        write_data_rows(ws17, 6, d17a, 3)
        top_count = len(d17a)
    else:
        ws17.cell(row=6, column=1, value="Aucun concept extrait (concepts_json absent ou vide)").font = Font(name="Arial", size=10, color="DC2626")
        top_count = 1

    # Section B : paires de concepts co-occurrents
    b17_start = 6 + top_count + 3
    ws17.cell(row=b17_start, column=1,
              value="B. Paires de concepts les plus souvent co-presents (meme article)").font = Font(name="Arial", bold=True, size=11)
    h17b = ["Rang", "Concept 1", "Concept 2", "Co-occurrences"]
    write_header_row(ws17, b17_start + 1, h17b, "7C3AED")
    if not concept_pairs_df.empty:
        d17b = [
            [rank + 1, r["concept_1"], r["concept_2"], int(r["co_occurrences"])]
            for rank, (_, r) in enumerate(concept_pairs_df.iterrows())
        ]
        write_data_rows(ws17, b17_start + 2, d17b, 4)
    else:
        ws17.cell(row=b17_start + 2, column=1, value="Aucune paire extraite").font = Font(name="Arial", size=10, color="DC2626")
    set_col_widths(ws17, {"A": 8, "B": 45, "C": 45, "D": 16})

    wb.save(output_path)
    print("Excel etape 3 : " + output_path)


# =====================================================================
# MAIN
# =====================================================================

def main():
    clean_path = "data/openalex_clean.parquet"
    excel_path = "outputs/etape3_analyse_descriptive.xlsx"

    print("=" * 70)
    print("  ETAPE 3 -- ANALYSE DESCRIPTIVE")
    print("=" * 70)

    if not os.path.exists(clean_path):
        print("ERREUR : " + clean_path + " introuvable.")
        print("  Lance d'abord : python 02_clean_and_prepare.py")
        return

    df = pd.read_parquet(clean_path)
    print(str(len(df)) + " articles charges")

    # Garde geo_zone meme si le parquet provient d'une ancienne version du script 02
    if "geo_zone" not in df.columns:
        print("  AVERTISSEMENT : colonne geo_zone absente. Relancez 02_clean_and_prepare.py.")
        df["geo_zone"] = "Unknown"

    print("\nCalcul des metriques de base...")
    metrics_q = compute_period_metrics(df, "quarter")
    metrics_y = compute_period_metrics(df, "year")

    print("Calcul prestige des revues...")
    df, journal_stats, p90, p75 = compute_prestige_tiers(df)
    prestige_y = compute_prestige_metrics_by_year(df)
    print("  Seuils : Top 10% >= " + str(round(p90, 1)) + " citations (med.) | Top 25% >= " + str(round(p75, 1)))

    print("Calcul metriques geographiques...")
    geo_y = compute_geo_metrics_by_year(df)

    print("Calcul percentile citations normalise par annee...")
    df = add_citation_percentile(df)

    print("Calcul Shannon ponderee (impact) vs volumique...")
    shannon_cmp = compute_shannon_weighted_by_year(df)

    print("Regression OLS Mega-Publisher...")
    ols_summary, coef_df = run_mega_publisher_regression(df, metrics_y)

    print("Calcul reseau de collaboration internationale...")
    country_cooc_matrix, country_counts, country_top_pairs = compute_country_cooccurrence(df)
    print("  " + str(len(country_counts)) + " pays distincts detectes")

    print("Calcul dynamique des equipes de recherche...")
    team_time, collab_impact, team_by_year = compute_team_dynamics(df)

    print("Calcul co-occurrence des concepts...")
    concept_top_df, concept_pairs_df = compute_concept_cooccurrence(df)
    print("  " + str(len(concept_top_df)) + " concepts uniques extraits")

    print("\nGeneration des figures...")
    fig01_volume(df)
    fig02_diversity(metrics_q)
    fig03_heatmap(df)
    fig04_lorenz(df)
    high_ai, low_ai = fig05_exposure(df)
    test_results = fig06_before_after(metrics_q)
    fig07_intensity(df)
    fig08_prestige(prestige_y)
    fig09_geographie(geo_y)
    fig_shannon_compare(shannon_cmp)
    fig_ols_coefs(coef_df)
    fig12_countries_network(country_cooc_matrix)
    fig13_team_dynamics(team_time, team_by_year, collab_impact)

    print("\nGeneration de l'Excel...")
    create_analysis_excel(df, metrics_q, metrics_y, test_results, high_ai, low_ai,
                          prestige_y, journal_stats, geo_y, shannon_cmp, ols_summary,
                          coef_df, country_cooc_matrix, country_top_pairs,
                          team_time, team_by_year, collab_impact,
                          concept_top_df, concept_pairs_df, excel_path)

    print()
    print("=" * 70)
    print("  ETAPE 3 TERMINEE")
    print("=" * 70)
    print("  Excel   : " + excel_path + " (17 onglets)")
    print("  Figures : " + FIG_DIR + "/ (13 PNG)")
    for f in sorted(os.listdir(FIG_DIR)):
        print("    " + f)


if __name__ == "__main__":
    main()