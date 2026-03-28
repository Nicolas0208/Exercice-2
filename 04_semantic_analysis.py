"""
=============================================================================
Script 04 -- Analyse semantique par embeddings (Version GPU & Checkpoints)
=============================================================================
Livrables :
  - outputs/etape4_semantique.xlsx   (2 onglets)
  - outputs/figures/fig10_semantic_diversity.png
  - outputs/embeddings_checkpoint.npy (Sauvegarde des calculs)

Methodologie :
  Chaque abstract est encode en vecteur dense (384 dims) via all-MiniLM-L6-v2.
  La diversite semantique annuelle est mesuree par la distance cosinus moyenne
  des articles au centroide de leur annee — approche O(N) en memoire, sans
  matrice N x N.
  Metrique complementaire : norme du centroide (proche de 0 = corpus diffus,
  proche de 1 = corpus domine par un seul theme).
=============================================================================
"""

# =====================================================================
# OPTIMISATION CPU — variables BLAS/OMP a fixer AVANT toute importation
# Ryzen 7 5700U : 8 coeurs / 16 threads, supporte AVX2
# =====================================================================
import os
_NCPU = os.cpu_count() or 8
os.environ.setdefault("OMP_NUM_THREADS",     str(_NCPU))
os.environ.setdefault("MKL_NUM_THREADS",     str(_NCPU))
os.environ.setdefault("OPENBLAS_NUM_THREADS",str(_NCPU))
os.environ.setdefault("NUMEXPR_NUM_THREADS", str(_NCPU))
os.environ.setdefault("TOKENIZERS_PARALLELISM", "true")

import numpy as np
import pandas as pd
import torch
torch.set_num_threads(_NCPU)          # PyTorch utilise tous les coeurs CPU
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import warnings
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
warnings.filterwarnings("ignore")

CLEAN_PATH = "data/openalex_clean.parquet"
EXCEL_PATH = "outputs/etape4_semantique.xlsx"
CSV_DISCIPLINE_PATH = "outputs/etape4_discipline_centroids.csv"
FIG_DIR = "outputs/figures"
CHECKPOINT_PATH = "outputs/embeddings_checkpoint.npy"
MODEL_NAME = "all-MiniLM-L6-v2"
# 256 = meilleur compromis debit / memoire sur CPU (etait 4096 → gaspillage)
BATCH_SIZE = 256
SAVE_INTERVAL = 100000
MIN_ABSTRACT_LEN = 100
MIN_ARTICLES_PER_DISC_YEAR = 5   # seuil min. pour calculer un centroide
MAX_DISCIPLINES = 12             # top-N disciplines les plus frequentes (lisibilite figures)

plt.rcParams.update({
    "figure.dpi": 150, "font.size": 11, "font.family": "sans-serif",
    "axes.spines.top": False, "axes.spines.right": False,
    "axes.grid": True, "grid.alpha": 0.3,
})

COLORS = {
    "primary": "#2563EB", "secondary": "#DC2626", "accent": "#059669",
    "gray": "#6B7280", "purple": "#7C3AED",
}


# =====================================================================
# EXCEL STYLES
# =====================================================================

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def write_header_row(ws, row, headers, color="2563EB"):
    fill = PatternFill("solid", fgColor=color)
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_data_rows(ws, start_row, data_rows, ncols):
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ALT_FILL


def set_col_widths(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def write_title(ws, row, text, color="2563EB"):
    ws.cell(row=row, column=1, value=text).font = Font(name="Arial", bold=True, size=14, color=color)


# =====================================================================
# CHARGEMENT DU MODELE — fastembed (ONNX INT8) ou sentence-transformers
# =====================================================================

def load_embedding_model(device):
    """
    Strategie de chargement par ordre de performance decroissante :
      1. fastembed  — ONNX INT8, 3-5x plus rapide sur CPU (pip install fastembed)
      2. sentence-transformers + quantisation dynamique INT8 via PyTorch (~2x)
      3. sentence-transformers standard (fallback)
    """
    # --- Option 1 : fastembed (ONNX quantise) ---
    # threads= controle intra_op_num_threads d'ONNX Runtime (independant de PyTorch)
    # Note Windows : necessite le mode Developpeur (droits symlinks) pour le cache HF.
    try:
        from fastembed import TextEmbedding
        try:
            fe_model = TextEmbedding("sentence-transformers/" + MODEL_NAME, threads=_NCPU)
        except TypeError:
            # Versions anterieures de fastembed sans le parametre threads
            fe_model = TextEmbedding("sentence-transformers/" + MODEL_NAME)
        print(f"  Backend : fastembed ONNX INT8 ({_NCPU} threads ONNX Runtime)")
        return ("fastembed", fe_model)
    except ImportError:
        pass
    except Exception as e:
        # Sur Windows sans mode Developpeur, les symlinks du cache HF echouent (WinError 1314)
        # -> fallback transparent vers sentence-transformers
        print(f"  fastembed indisponible ({type(e).__name__}), fallback sentence-transformers")

    # --- Options 2 & 3 : sentence-transformers ---
    from sentence_transformers import SentenceTransformer
    model = SentenceTransformer(MODEL_NAME, device=device)

    if device == "cpu":
        # Quantisation dynamique INT8 — exploite les instructions AVX2 du 5700U
        # torch.ao.quantization est l'API courante (PyTorch >= 1.13)
        try:
            import torch.ao.quantization as _tq
            _quantize_fn = getattr(_tq, "quantize_dynamic")
            model._modules["0"].auto_model = _quantize_fn(
                model._modules["0"].auto_model,
                {torch.nn.Linear},
                dtype=torch.qint8,
            )
            print("  Backend : sentence-transformers + quantisation INT8 (~2x sur CPU)")
            return ("st", model)
        except Exception as e:
            print(f"  Quantisation INT8 non disponible ({e}), fallback float32")

    print("  Backend : sentence-transformers float32")
    return ("st", model)


# =====================================================================
# EMBEDDINGS AVEC CHECKPOINTS
# =====================================================================

def generate_embeddings(model_tuple, abstracts):
    """
    Encode par tranches de SAVE_INTERVAL avec tri par longueur (reduit le padding).
    Supporte fastembed (ONNX) et sentence-transformers.
    Resume automatiquement depuis le dernier checkpoint.
    """
    backend, model = model_tuple
    n = len(abstracts)
    start_idx = 0
    all_emb = []

    if os.path.exists(CHECKPOINT_PATH):
        try:
            saved_emb = np.load(CHECKPOINT_PATH)
            start_idx = len(saved_emb)
            all_emb = [saved_emb]
            print(f"  Reprise a l'index {start_idx} / {n}")
            if start_idx >= n:
                return saved_emb
        except Exception as e:
            print(f"  Erreur checkpoint : {e}. Reprise a zero.")
            start_idx = 0
            all_emb = []

    for chunk_start in range(start_idx, n, SAVE_INTERVAL):
        chunk_end = min(chunk_start + SAVE_INTERVAL, n)
        chunk = abstracts[chunk_start:chunk_end]
        print(f"  Encodage tranche {chunk_start}-{chunk_end} / {n}...")
        t0 = time.time()

        if backend == "fastembed":
            # ONNX Runtime tire plus parti des grands batches qu'un appel encode() PyTorch.
            # On utilise 512 au lieu de BATCH_SIZE (256) pour reduire l'overhead Python.
            fe_batch = min(512, len(chunk))
            emb = np.array(
                list(model.embed(chunk, batch_size=fe_batch)),
                dtype=np.float32,
            )
        else:
            emb = model.encode(
                chunk,
                batch_size=BATCH_SIZE,
                show_progress_bar=True,
                convert_to_numpy=True,
                normalize_embeddings=True,
                sort_by_length=True,     # minimise le padding -> +20-30% vitesse
            ).astype(np.float32)

        elapsed = time.time() - t0
        rate = (chunk_end - chunk_start) / max(elapsed, 1e-6)
        print(f"  [SAUVEGARDE] {chunk_end} / {n}  ({rate:.0f} abstracts/s)")

        all_emb.append(emb)
        np.save(CHECKPOINT_PATH, np.concatenate(all_emb, axis=0))

    return np.concatenate(all_emb, axis=0)


# =====================================================================
# DIVERSITE SEMANTIQUE (centroide)
# =====================================================================

def semantic_diversity(E):
    """Distance cosinus moyenne au centroide (E doit etre L2-normalise)."""
    n = len(E)
    if n < 2:
        return {"mean_cosine_dist": np.nan, "std_cosine_dist": np.nan, "centroid_norm": np.nan}

    centroid = E.mean(axis=0)
    c_norm = float(np.linalg.norm(centroid))
    c_unit = centroid / c_norm if c_norm > 0 else centroid
    cos_dist = 1.0 - (E @ c_unit)

    return {
        "mean_cosine_dist": float(np.mean(cos_dist)),
        "std_cosine_dist": float(np.std(cos_dist)),
        "centroid_norm": c_norm,
    }


# =====================================================================
# ANALYSE INTRA / INTER-CENTROIDES PAR DISCIPLINE
# =====================================================================

def compute_centroid(E):
    """E doit etre L2-normalise."""
    return E.mean(axis=0)


def cosine_dist(a, b):
    na = np.linalg.norm(a)
    nb = np.linalg.norm(b)
    if na == 0 or nb == 0:
        return np.nan
    return float(1.0 - np.dot(a, b) / (na * nb))


def compute_discipline_centroid_metrics(E):
    """E doit etre L2-normalise."""
    n = len(E)
    if n < MIN_ARTICLES_PER_DISC_YEAR:
        return None
    centroid = E.mean(axis=0)
    c_norm = np.linalg.norm(centroid)
    c_unit = centroid / c_norm if c_norm > 0 else centroid
    cos_dist = 1.0 - (E @ c_unit)
    return {
        "centroid": centroid,
        "mean_intra_dist": float(np.mean(cos_dist)),
        "std_intra_dist": float(np.std(cos_dist)),
        "centroid_norm": float(c_norm),
        "n": n,
    }


def compute_all_discipline_metrics(df, E_all, top_disciplines):
    """Utilise groupby pour precomputer les indices en O(N) au lieu de
    masques booleens repetes O(N x nb_annees)."""
    top_disc_set = set(top_disciplines)
    records = []
    prev_centroids = {}

    # O(N) unique : tous les groupes (year, discipline) indexes d'un coup
    for (year, disc), group_df in df.groupby(["year", "primary_discipline"]):
        if disc not in top_disc_set:
            continue
        indices = group_df.index.values
        if len(indices) < MIN_ARTICLES_PER_DISC_YEAR:
            prev_centroids.pop(disc, None)
            continue

        m = compute_discipline_centroid_metrics(E_all[indices])
        if m is None:
            continue

        inter_shift = np.nan
        if disc in prev_centroids:
            inter_shift = cosine_dist(m["centroid"], prev_centroids[disc])
        prev_centroids[disc] = m["centroid"]

        records.append({
            "discipline": disc,
            "year": int(year),
            "n_articles": m["n"],
            "mean_intra_dist": round(m["mean_intra_dist"], 6),
            "std_intra_dist": round(m["std_intra_dist"], 6),
            "centroid_norm": round(m["centroid_norm"], 6),
            "inter_shift": round(float(inter_shift), 6) if not np.isnan(inter_shift) else np.nan,
        })

    return pd.DataFrame(records).sort_values(["discipline", "year"]).reset_index(drop=True)


def fig_discipline_intra(disc_df, top_disciplines):
    fig, ax = plt.subplots(figsize=(14, 6))
    cmap = plt.cm.tab20(np.linspace(0, 1, len(top_disciplines)))

    for disc, color in zip(top_disciplines, cmap):
        sub = disc_df[disc_df["discipline"] == disc].sort_values("year")
        if sub.empty:
            continue
        label = disc[:25] + "..." if len(disc) > 25 else disc
        ax.plot(sub["year"], sub["mean_intra_dist"], "o-", color=color,
                lw=1.5, ms=4, alpha=0.85, label=label)

    ax.axvline(2022.5, color="gray", ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee")
    ax.set_ylabel("Distance cosinus moyenne au centroide (diversite intra)")
    ax.set_title("Variance semantique intra-discipline par annee\n"
                 "(distance cosinus moyenne articles -> centroide C_{d,t})")
    ax.legend(fontsize=7, ncol=2, loc="upper left")
    plt.tight_layout()
    path = FIG_DIR + "/fig12_intra_centroid.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    print("  Figure 12 : Variance intra-discipline")
    return path


def fig_discipline_inter_shift(disc_df, top_disciplines):
    fig, ax = plt.subplots(figsize=(14, 6))
    cmap = plt.cm.tab20(np.linspace(0, 1, len(top_disciplines)))

    for disc, color in zip(top_disciplines, cmap):
        sub = disc_df[(disc_df["discipline"] == disc) & disc_df["inter_shift"].notna()].sort_values("year")
        if sub.empty:
            continue
        label = disc[:25] + "..." if len(disc) > 25 else disc
        ax.plot(sub["year"], sub["inter_shift"], "s--", color=color,
                lw=1.5, ms=4, alpha=0.85, label=label)

    ax.axvline(2022.5, color="gray", ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee")
    ax.set_ylabel("Distance cosinus C_{d,t-1} -> C_{d,t}  (shift paradigmatique)")
    ax.set_title("Saut paradigmatique inter-centroide par discipline\n"
                 "(distance entre centroide annee t et annee t-1)")
    ax.legend(fontsize=7, ncol=2, loc="upper left")
    plt.tight_layout()
    path = FIG_DIR + "/fig13_inter_shift.png"
    plt.savefig(path, bbox_inches="tight")
    plt.close()
    print("  Figure 13 : Shift inter-centroide")
    return path


# =====================================================================
# FIGURE
# =====================================================================

def fig10_semantic(sem_df):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    ax = axes[0]
    ax.plot(sem_df["year"], sem_df["semantic_diversity"], "o-",
            color=COLORS["primary"], lw=2, ms=6, label="Tous articles")
    ax.fill_between(
        sem_df["year"],
        sem_df["semantic_diversity"] - sem_df["semantic_std"],
        sem_df["semantic_diversity"] + sem_df["semantic_std"],
        alpha=0.15, color=COLORS["primary"],
    )
    if sem_df["sem_div_ai"].notna().any():
        ax.plot(sem_df["year"], sem_df["sem_div_ai"], "s--",
                color=COLORS["secondary"], lw=1.5, ms=5, label="Avec mention IA")
    if sem_df["sem_div_nonai"].notna().any():
        ax.plot(sem_df["year"], sem_df["sem_div_nonai"], "^--",
                color=COLORS["accent"], lw=1.5, ms=5, label="Sans mention IA")
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5, label="ChatGPT")
    ax.set_xlabel("Annee")
    ax.set_ylabel("Distance cosinus moyenne au centroide")
    ax.set_title("(a) Diversite semantique par annee")
    ax.legend(fontsize=9)

    ax = axes[1]
    ax.plot(sem_df["year"], sem_df["centroid_norm"], "o-",
            color=COLORS["purple"], lw=2, ms=6)
    ax.axvline(2022.5, color=COLORS["gray"], ls="--", alpha=0.5)
    ax.set_xlabel("Annee")
    ax.set_ylabel("Norme du centroide  (0 = diffus  /  1 = concentre)")
    ax.set_title("(b) Concentration semantique (norme centroide)")

    plt.suptitle(
        "Diversite semantique des abstracts par embeddings (" + MODEL_NAME + ")",
        fontsize=12, fontweight="bold", y=1.02,
    )
    plt.tight_layout()
    plt.savefig(FIG_DIR + "/fig10_semantic_diversity.png", bbox_inches="tight")
    plt.close()
    print("  Figure 10 : Diversite semantique")


# =====================================================================
# EXCEL
# =====================================================================

def create_semantic_excel(sem_df, disc_df, output_path):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "1. Diversite semantique"
    ws1.sheet_properties.tabColor = "2563EB"
    write_title(ws1, 1, "DIVERSITE SEMANTIQUE PAR ANNEE (embeddings " + MODEL_NAME + ")")
    ws1.cell(row=2, column=1, value="Date : " + datetime.now().strftime("%Y-%m-%d %H:%M")).font = Font(name="Arial", size=10, italic=True, color="6B7280")

    h1 = ["Annee", "N articles", "N mention IA",
          "Diversite sem. (tous)", "Ecart-type", "Norme centroide",
          "Diversite sem. (IA)", "Diversite sem. (non-IA)"]
    write_header_row(ws1, 4, h1)

    d1 = []
    for _, r in sem_df.iterrows():
        d1.append([
            int(r["year"]), int(r["n_articles"]), int(r["n_ai"]),
            round(r["semantic_diversity"], 6) if pd.notna(r["semantic_diversity"]) else "N/A",
            round(r["semantic_std"], 6) if pd.notna(r["semantic_std"]) else "N/A",
            round(r["centroid_norm"], 6) if pd.notna(r["centroid_norm"]) else "N/A",
            round(r["sem_div_ai"], 6) if pd.notna(r["sem_div_ai"]) else "N/A",
            round(r["sem_div_nonai"], 6) if pd.notna(r["sem_div_nonai"]) else "N/A",
        ])
    write_data_rows(ws1, 5, d1, 8)
    set_col_widths(ws1, {
        "A": 10, "B": 12, "C": 14, "D": 22, "E": 14, "F": 18, "G": 20, "H": 22,
    })

    ws2 = wb.create_sheet("2. Centroides disciplines")
    ws2.sheet_properties.tabColor = "059669"
    write_title(ws2, 1, "ANALYSE INTRA / INTER-CENTROIDES PAR DISCIPLINE (" + MODEL_NAME + ")")
    ws2.cell(row=2, column=1,
             value="mean_intra_dist : variance semantique interne | inter_shift : saut paradigmatique C_{d,t-1}->C_{d,t}").font = Font(name="Arial", size=9, italic=True, color="6B7280")

    h2 = ["Discipline", "Annee", "N articles",
          "Dist. intra moy.", "Dist. intra ET", "Norme centroide",
          "Shift inter-annee"]
    write_header_row(ws2, 4, h2, "059669")

    d2 = []
    for _, r in disc_df.iterrows():
        shift_val = round(float(r["inter_shift"]), 6) if pd.notna(r["inter_shift"]) else "N/A"
        d2.append([
            str(r["discipline"]),
            int(r["year"]),
            int(r["n_articles"]),
            round(r["mean_intra_dist"], 6),
            round(r["std_intra_dist"], 6),
            round(r["centroid_norm"], 6),
            shift_val,
        ])
    write_data_rows(ws2, 5, d2, 7)
    set_col_widths(ws2, {
        "A": 30, "B": 10, "C": 12, "D": 18, "E": 16, "F": 18, "G": 18,
    })

    ws2 = wb.create_sheet("3. Methodologie")
    ws2.sheet_properties.tabColor = "6B7280"
    write_title(ws2, 1, "NOTES METHODOLOGIQUES - ANALYSE SEMANTIQUE", color="6B7280")

    notes = [
        ["Modele", MODEL_NAME + " (sentence-transformers, ~80 Mo, 384 dimensions)"],
        ["Taille batch", str(BATCH_SIZE) + " abstracts par appel model.encode()"],
        ["Filtre abstracts", "Longueur >= " + str(MIN_ABSTRACT_LEN) + " caracteres"],
        ["Normalisation", "Chaque embedding normalise L2 avant calcul du centroide"],
        ["Centroide", "Moyenne des embeddings normalises du groupe annuel"],
        ["Distance cosinus", "1 - (E_i . centroide / ||centroide||)  -- entre 0 et 1"],
        ["mean_cosine_dist", "Diversite semantique : 0=articles identiques, 1=corpus totalement disperse"],
        ["centroid_norm", "Norme L2 du centroide : 0=diffus (themes multiples), 1=concentre (un theme domine)"],
        ["Avantage memoire", "Complexite O(N*D) au lieu de O(N^2) pour les distances paires"],
        ["Sous-groupes", "IA vs non-IA : min. 5 articles par groupe requis"],
        ["Interpretation", "Une baisse de mean_cosine_dist apres 2022 = homogeneisation semantique du corpus"],
    ]

    h2 = ["Concept", "Description"]
    write_header_row(ws2, 3, h2, "6B7280")
    write_data_rows(ws2, 4, notes, 2)
    for i in range(len(notes)):
        ws2.cell(row=4 + i, column=1).font = Font(name="Arial", size=10, bold=True)
        ws2.cell(row=4 + i, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    set_col_widths(ws2, {"A": 22, "B": 80})

    wb.save(output_path)
    print("Excel etape 4 : " + output_path)


# =====================================================================
# MAIN
# =====================================================================

def main():
    t_start = time.time()
    print("=" * 70)
    print("  ETAPE 4 -- ANALYSE SEMANTIQUE PAR EMBEDDINGS (V3 CPU-optimise)")
    print("=" * 70)
    print(f"  Modele : {MODEL_NAME}  |  Threads CPU : {_NCPU}  |  Batch : {BATCH_SIZE}")

    if not os.path.exists(CLEAN_PATH):
        print("ERREUR : " + CLEAN_PATH + " introuvable.")
        print("  Lancez d'abord : python 02_clean_and_prepare.py")
        return

    os.makedirs(FIG_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

    print("\nChargement des donnees...")
    df = pd.read_parquet(
        CLEAN_PATH,
        columns=["openalex_id", "abstract", "year", "ai_mention", "primary_discipline"],
    )
    df = df[df["abstract"].str.len() >= MIN_ABSTRACT_LEN].copy().reset_index(drop=True)
    print("  " + str(len(df)) + " articles avec abstract suffisant")

    top_disciplines = df["primary_discipline"].value_counts().head(MAX_DISCIPLINES).index.tolist()

    print("\nChargement du modele sentence-transformers...")

    # Detection du materiel (GPU / Apple Silicon / CPU)
    if torch.cuda.is_available():
        device = "cuda"
    elif torch.backends.mps.is_available():
        device = "mps"
    else:
        device = "cpu"
    print("  Appareil detecte : " + device.upper())

    model_tuple = load_embedding_model(device)

    print("\nEncodage global des " + str(len(df)) + " abstracts...")
    all_embeddings = generate_embeddings(model_tuple, df["abstract"].tolist())
    print("  Embeddings calcules : shape " + str(all_embeddings.shape))

    # normalize_embeddings=True dans model.encode garantit deja la normalisation L2.
    # On renormalise quand meme pour securiser la reprise depuis un vieux checkpoint.
    norms = np.linalg.norm(all_embeddings, axis=1, keepdims=True)
    norms = np.where(norms == 0, 1.0, norms)
    E_all = (all_embeddings / norms).astype(np.float32)
    del all_embeddings  # libere ~2 GB RAM

    results = []
    # groupby precompute les indices une seule fois en O(N)
    for year, year_df in df.groupby("year"):
        year_indices = year_df.index.values
        year_emb = E_all[year_indices]
        ai_mask = year_df["ai_mention"].values.astype(bool)
        n = len(year_indices)
        print("\nAnnee " + str(year) + " (" + str(n) + " articles)...")

        metrics = semantic_diversity(year_emb)
        ai_emb = year_emb[ai_mask]
        nonai_emb = year_emb[~ai_mask]
        ai_m = semantic_diversity(ai_emb) if len(ai_emb) >= 5 else {}
        nonai_m = semantic_diversity(nonai_emb) if len(nonai_emb) >= 5 else {}

        results.append({
            "year": int(year),
            "n_articles": n,
            "n_ai": int(ai_mask.sum()),
            "semantic_diversity": metrics["mean_cosine_dist"],
            "semantic_std": metrics["std_cosine_dist"],
            "centroid_norm": metrics["centroid_norm"],
            "sem_div_ai": ai_m.get("mean_cosine_dist", np.nan),
            "sem_div_nonai": nonai_m.get("mean_cosine_dist", np.nan),
        })
        print("  -> diversite semantique = " + str(round(metrics["mean_cosine_dist"], 4)))

    sem_df = pd.DataFrame(results)

    print("\nCalcul des centroides par discipline et annee...")
    disc_df = compute_all_discipline_metrics(df, E_all, top_disciplines)

    os.makedirs(os.path.dirname(CSV_DISCIPLINE_PATH), exist_ok=True)
    disc_df.to_csv(CSV_DISCIPLINE_PATH, index=False, encoding="utf-8-sig")

    print("\nGeneration des figures...")
    fig10_semantic(sem_df)
    if not disc_df.empty:
        fig_discipline_intra(disc_df, top_disciplines)
        fig_discipline_inter_shift(disc_df, top_disciplines)

    print("Generation de l'Excel...")
    create_semantic_excel(sem_df, disc_df, EXCEL_PATH)

    elapsed_total = time.time() - t_start
    print(f"\n  ETAPE 4 TERMINEE en {elapsed_total:.1f}s ({elapsed_total/60:.1f} min)")


if __name__ == "__main__":
    main()
